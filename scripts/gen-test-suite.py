#!/usr/bin/env python
"""Generate the DoubleDone end-to-end manual test suite.

Single source of truth for the manual QA pass. Emits:
  - docs/qa/DoubleDone-E2E-Test-Suite.xlsx   (fillable: Result dropdown + Findings + Date)
  - docs/qa/e2e-test-suite.md                (readable / diffable in git)

Run:  python scripts/gen-test-suite.py
Deps: openpyxl  (python -m pip install openpyxl)

Edit CASES below to change the suite, then re-run. The .xlsx is a TEMPLATE: copy
it before a run, or fill it in place and keep your copy outside git.
"""

from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Dusk palette, so the sheet feels like the app.
ACCENT = "9B6A7D"
INK = "2B2722"
LINE = "ECE4D8"
BG = "FAF6F1"
PASS_FILL = "DDEAD0"
FAIL_FILL = "F2D7DE"
BLOCK_FILL = "F4E7C9"

# Each case: (id, area, priority, test, steps, expected, platform)
CASES = [
    # --- Capture & brain-dump -------------------------------------------------
    ("CAP-01", "Capture", "P1", "Add a single task",
     "On Today, tap the capture box, type 'Buy milk', submit.",
     "The task appears in Today immediately. No reload needed.", "Both"),
    ("CAP-02", "Capture", "P1", "Brain-dump several tasks",
     "Enter several tasks in a row (e.g. 5 quick ones).",
     "Each becomes its own task in order. None lost or merged.", "Both"),
    ("CAP-03", "Capture", "P2", "One-off future date (door > Pick a date)",
     "Type a task, open the 'WHEN · REPEATING · STEPS' door, tap 'Pick a date', choose a day next week, then Add.",
     "The Add button reads 'Add · <that day>' before the tap. Task is scheduled for that date and does NOT show in Today until then.", "Both"),
    ("CAP-04", "Capture", "P2", "The when chips (door > Today / Tomorrow)",
     "Type a task, open the door, pick Tomorrow, Add. Repeat with Today.",
     "The door summary and the Add label name the chosen day ('Add · Tomorrow'). Task lands on that day; Today shows only today's.", "Both"),
    ("CAP-05", "Capture", "P3", "Empty / whitespace capture",
     "Submit an empty box, then a box of only spaces.",
     "Nothing is added. No error, no blank row.", "Both"),
    ("CAP-06", "Capture", "P3", "Very long title",
     "Capture a task with a very long title (a full sentence+).",
     "Wraps or truncates gracefully. No layout break or overflow.", "Both"),
    ("CAP-07", "Capture", "P2", "Talk-to-capture: speak a brain-dump (web)",
     "On the web app in Chrome / Edge / Safari, tap the '🎤 Speak' button under the capture box and allow the mic. Say a few tasks, pausing between each, then tap 'Listening… tap to stop'.",
     "Each spoken phrase lands on its own line; two or more lines surface 'Sort for me', which sorts them as normal. The browser does the speech-to-text (no audio reaches our servers), so only text leaves the device, and only if you Sort. Tapping stop returns the button to 'Speak' and keeps the captured lines.", "Web"),
    ("CAP-08", "Capture", "P3", "Talk-to-capture hides where unsupported",
     "Open the web app in Firefox (no Web Speech API), and separately open the Android app.",
     "The '🎤 Speak' button is simply absent (no error, no setting) on Firefox and on native Android, where the Gboard keyboard mic already dictates into the box.", "Both"),
    ("CAP-09", "Capture", "P2", "Tidy a run-on or rambly line into tasks (AI)",
     "Type or speak a single long line, either several things ('buy milk and walk the dog and email Sarah') or one rambly thought ('I feel like I want to do something fun'), then tap 'Tidy this into tasks'.",
     "The line is replaced by clean task(s) in your own words, nothing invented or reordered. A run-on becomes several lines (and 'Sort for me' appears); a single rambly thought becomes one tidy task ('do something fun'). On failure it degrades calmly ('Couldn't tidy that just now'), text kept. Works on web and Android. Needs the Worker deployed with /split.", "Both"),
    ("CAP-10", "Capture", "P1", "The one door: weekly end to end",
     "Open capture ('+ Add to…'), type 'water the plants', open the 'WHEN · REPEATING · STEPS' door, tap Weekly, then Add.",
     "The panel opens with the input focused (typing works instantly). The door summary reads 'Today · Weekly on <today's day>' and the Add button reads 'Add · Weekly on <day>' BEFORE the tap. After Add the task lands with the repeat glyph and the door resets to 'Today' with the composer closed. Weekly is 3 taps beyond typing, never more.", "Both"),
    ("CAP-11", "Capture", "P1", "Text is never lost; the frame yields",
     "Open capture, type half a thought, tap Close. Then reopen capture.",
     "While capture is open the 'Right now' action slot is away; it returns the moment capture closes. Reopening shows the typed text still there, input focused, with the door reset to 'Today · no repeat · no steps'. Nothing typed is ever lost by a collapse.", "Both"),
    ("CAP-13", "Capture", "P1", "The open panel rides ABOVE the keyboard, never under it",
     "On a real Android phone (app and browser separately), tap '+ Add to…' and watch where the panel lands as the keyboard rises. Type a line, open the door, pick Weekly, Add. Rotate through a large font-size setting and repeat.",
     "The whole panel (input, door, Add) sits above the keyboard at all times; you can see what you type and reach Add without dismissing the keyboard. The inscription and footer links tuck away while the keyboard is up and return when it goes. Opening the door tucks the keyboard and shows the composer. REGRESSION GUARD: a tester's v19 screenshots showed only 'ADD / Close' with everything else under the keyboard, because edge-to-edge Android ignores adjustResize and Chrome 108+ overlays the keyboard by default; the fix is the keyboard-height footer lift (native) + interactive-widget=resizes-content (web).", "Android"),
    ("CAP-12", "Capture", "P3", "Steps go quiet, never vanish",
     "Type several lines in capture and open the door. Then clear to one line and pick Weekly.",
     "The Steps row keeps its place both times, at lowered contrast, saying why in plain words ('Steps are for one thing at a time.' / 'Steps are for tasks that do not repeat.'). It never disappears and never errors. With one line and no repeat the stepper works and the count joins the door summary ('Today · 3 steps').", "Both"),

    # --- Today & the daily loop ----------------------------------------------
    ("TOD-01", "Today", "P1", "Complete a task",
     "Tap a task's done control on Today.",
     "Marked done with calm feedback. No shame language anywhere.", "Both"),
    ("TOD-02", "Today", "P1", "Today is sized to be doable",
     "Add many tasks across days; look at Today.",
     "Today shows today's achievable set, not the entire backlog.", "Both"),
    ("TOD-03", "Today", "P1", "Push a task to tomorrow",
     "Use the Tomorrow action on a task in Today.",
     "Leaves Today and appears tomorrow. No guilt framing.", "Both"),
    ("TOD-04", "Today", "P1", "Close the day -> rested state",
     "Tap 'Close the day', read the wrap, optionally type a final win in 'Anything else you did?', then tap 'Goodnight'.",
     "Wrap celebrates what you finished (zero guilt, unfinished never shamed). Any text in 'Anything else you did?' is logged as a completed task. After Goodnight, Today becomes a calm 'You've closed today' rested screen with the task list + capture hidden. Survives reload.", "Both"),
    ("TOD-04b", "Today", "P2", "Reopen a closed day",
     "From the rested screen, tap 'Reopen today'.",
     "Returns to the normal Today with all tasks intact. (A new calendar day also auto-clears the closed state.)", "Both"),
    ("TOD-04c", "Today", "P2", "Add for tomorrow after closing the day",
     "Close the day to reach the rested screen. Under the rest message, type a thought in the 'Something on your mind for tomorrow?' box and tap 'Add for tomorrow'. Then test 'Not tomorrow?' to pick a later day.",
     "Saved with a calm 'Saved for tomorrow. Rest well.' (or the picked day). The day STAYS closed (never reopens) and the new task does NOT show on the closed today. It is waiting on Today the next day. No AI, nothing leaves the device.", "Both"),
    ("TOD-04d", "Today", "P2", "Close a very full day (16+ finished) without getting stuck",
     "Finish a large number of tasks (roughly 16 or more) so the close-the-day wrap has a long finished list, then tap 'Close the day'.",
     "The wrap fits the screen: the finished list scrolls inside a bounded area while the roll-forward line, the 'Anything else you did?' note, and the 'Goodnight' button all stay visible and reachable. Goodnight is never pushed off-screen no matter how many things got done. (Regression: a long finished day used to hide Goodnight with no way to proceed.)", "Both"),
    ("TOD-05", "Today", "P2", "Undo a completion",
     "Complete a task, then undo it (toggle back).",
     "Returns to open cleanly. Counts/Calendar stay consistent.", "Both"),
    ("TOD-06", "Today", "P1", "Persistence across restart",
     "Add tasks, fully close the app/tab, reopen.",
     "Tasks are still there (local-first). Nothing lost.", "Both"),
    ("TOD-07", "Today", "P1", "Tap-and-hold a task -> its own actions, in place",
     "SCROLL THE LIST DOWN FIRST (this is the regression trigger), then press and hold a task. Do it on BOTH Standard and Quiet appearances, and on a short day (2 tasks) as well as a long one.",
     "The page does NOT jump: the scroll position stays exactly where it was. The screen does NOT flip into checkboxes, and '+ I also did that' plus the day actions stay visible. The held row expands IN PLACE into its own card in the DESIGN 1a shape (2026-07-25): the title (tappable to rename, faint underline), then the LEAD actions each on their own full-width row - Break it down (the tinted HERO: a solid accent fill with white label in light, a soft tint with accent label in dark, sub-label 'into small steps') / Make it tiny ('the first step') / Move to... / Mark as a lot (sub 'weight', which tints to an active state with a check when on) - then a 'More' row (caret + a faint 'Steps, Pin...' preview) that, tapped, reveals Steps (with a '2 of 5' when the task is in steps) / Pin / Remind me (native only) in place and flips its caret; then a hairline above the way out: Close (accent, bottom-LEFT, easiest thumb reach) / Select more / Remove (muted brick, far RIGHT, out of the reflex path). Fewer visible actions than the old flat grid (4 vs 11), same feature set. On a REPEATING task, Remove reads 'Skip today' (the series continues, never a delete). Every label renders in full with no clipped glyph even at large system font (each action is a full-width row, never a tight equal-width column). Both appearances behave identically; in Quiet the hero drops its fill for accent text held by whitespace.", "Both"),
    ("TOD-07b", "Today", "P2", "First-run coachmark teaches the long-press",
     "On a device that has never seen it (or after clearing 'doubledone.holdhint.v1'), open Today with at least one task. Read the hint, tap 'Got it', then reload. Separately check an empty Today and the rested (closed-day) screen.",
     "A calm one-line hint sits above the list: 'Hold a task for more: break it down, make it tiny, move it, or mark it as a lot.' (design v2 card, 2026-08-08: those four lead, then the reorder rail, then the More fold whose preview honestly names its everyday contents, 'Steps · Remind me', with Pin LAST inside carrying a honey ✦ and the sub 'holds the top': premium recedes, never advertises; 'Undo a step' dims in place with 'no steps yet' rather than vanishing. The card ends on the SHELF: a quiet tinted band rounded into the card's bottom corners holding Close (left), Select more, Remove (far right). The card rises in over ~180ms with a small settle; under reduce-motion it dissolves in ~90ms with no movement. The done-task card puts its 'a finished thing needs nothing shaped' inscription ABOVE the shelf.) 'Got it' dismisses the hint for good (survives reload). It never shows on an empty Today or the rested screen, and never returns once dismissed.", "Both"),
    ("TOD-29", "Today", "P2", "A held card opening near the bottom of the screen stays whole (web)",
     "On web with 8+ tasks, scroll so a task sits in the lower third of the window, then hold it.",
     "The view nudges just enough that the WHOLE card, including the Close row at its bottom, is visible (no smooth animation, one small instant correction; nothing moves when the card already fits). Before this fix the card's lower half, More and the way out included, opened below the fold and read as 'the card has a flat edge' (field report 2026-08-08). Native behaviour unchanged for now; its pass rides the held-card redesign.", "Web"),
    ("ORD-01", "Today", "P2", "Free manual reorder from the held card (Move up / Move down)",
     "With 4+ open tasks, hold one mid-list. Tap 'Move up' twice, watch it travel, then Close. Reload the app. Also check: the FIRST task ('Move up' dimmed), the LAST ('Move down' dimmed), a DONE task (no reorder row at all), and with a premium pin at the top ('Move up' on the second task is dimmed, and the pinned task itself has no reorder row).",
     "The held card shows THE RAIL (design v2): '↑ Move up | ↓ Move down' as one segmented hairline-bordered control, the card's ONLY bordered element (bordered = act-and-stay), FREE for everyone. Each tap moves the task ONE place, the card STAYS OPEN, and a screen reader hears 'Moved up, {n} of {m}'. The new order survives reload (manualOrder, local-only). An edge dims its cell in place; a PINNED task shows the whole rail rested (both cells dimmed, the pin holds the top); a DONE task's card has no rail at all. Nothing can be moved above a pinned task.", "Both"),
    ("TOD-07c", "Today", "P2", "Split a task into steps (and resize or make whole)",
     "Tap-and-hold a single task, then tap 'Steps' in the held row's actions. With the - / + stepper pick a number of parts (2-50) and tap Done. Tap the task on Today to advance a step. Hold it again, 'Steps', change the count, Done. Then hold, 'Steps', 'Make it whole again'.",
     "'Steps' opens a 'Track in steps' editor with a stepper bounded at 2 and 50. Done turns the task into an N-part one with a step counter; tapping it on Today advances one step, and finishing the last step completes it (with the whole-task bloom). Re-opening lets you change N: progress carries over, and shrinking below what's already done snaps it down. 'Make it whole again' drops the parts back to one task, keeping its done state. Single-task, non-recurring, not-done only. Free for everyone. REGRESSION CHECK: holding an ALREADY-SLICED task must show its full card (title with the live n / N count, Undo a step, Steps, Make it tiny, Move to..., Pin, Mark as a lot, Select more, Remove, Close) and NOT Break down (it is already in steps). A sliced task used to get a stripped-down hold, so this is the case that proves nothing was lost.", "Both"),
    ("TOD-08", "Today", "P2", "Shame-free re-entry after a gap",
     "Simulate not opening the app for 4+ days (set localStorage 'doubledone.lastopen.v1' to a date 5+ days ago), then reload Today.",
     "A calm 'Welcome back, the past is fine, here's just today' card appears above Today, never '47 overdue'. 'Start fresh' dismisses it; reopening same-day does not re-show it.", "Both"),
    ("TOD-09", "Today", "P2", "Log an off-list thing you did",
     "Tap '+ I also did that' (beneath the task list), type something you did, then Add it.",
     "It appears checked on Today and in the Calendar for today, counted as a completion, never as an unfinished task.", "Both"),
    ("TOD-10", "Today", "P2", "Focus mode: pick-and-go",
     "Tap 'Focus on one thing' (the prominent entry above the list). On 'Which one?' pick a task; try 'Done' (complete), 'Choose another' (back to the list), and Exit. THEN repeat with a LONG Today (15+ open tasks) and scroll the picker to its very last row.",
     "Full-screen single task, everything else hidden. Done completes it and returns to 'Which one?'; Choose another returns without completing; Exit closes. When none left: 'That's everything for now.'. With a LONG Today the picker SCROLLS and every row is reachable, including the last, while Exit stays pinned in its corner and never scrolls away. EXIT MUST ALSO BE CLICKABLE, especially by mouse on web: the scroll fix originally rendered the full-bleed scroller AFTER the absolutely-positioned Exit, so the scroller stacked ON TOP of it and every click landed on the scroll surface instead; Focus could not be left by mouse at all (Melroy, web, 2026-07-25). The Exit now renders after the scroller with a zIndex, so it always wins the stack. Regression check: click Exit dead-centre with a mouse, on web, on a day with 6+ tasks. Regression guard: the picker used to be a centred non-scrolling column, so a list taller than the screen was clipped at BOTH ends and those tasks could not be selected at all (found 2026-07-25).", "Both"),
    ("TOD-11", "Today", "P3", "Weight-of-today gauge",
     "Add a few tasks and watch the slim gauge under 'Just today'.",
     "A calm bar + warm label ('A gentle day. Room to breathe.' up to ~4, then 'A full day, but doable.', then 'A lot on. Be gentle with yourself.') reflects the count of unfinished one-offs, honest, never alarming, hidden on a clear day.", "Both"),
    ("TOD-12", "Today", "P2", "Multi-select bulk actions (reached deliberately, via Select more)",
     "Hold a task, tap 'Select more' to enter selection with it already ticked, tap more rows to add them (or 'Select all'), then use the bar. Cancel exits.",
     "Rows become checkboxes (circles cross-fade in; selected fills accent with a tick) and the count updates ('{n} selected', announced). The bar is now THE SELECT SHELF (congruency pass 2026-08-08): a card in the held-card-v2 family that RISES in like the card. Fixed anatomy in EVERY state, so it never changes height mid-selection: (1) count + Select all as a bordered pill that DIMS in place once everything is selected; (2) the verb row, Done / Move to... / Mark as a lot (reads 'Not a lot' only when every selected task is marked), dimming whole when nothing is selected; (3) Combine's PERMANENTLY RESERVED slot, where the surface's one tinted hero ('Combine' + 'merges similar tasks') fades in at 2+ combinable and announces its arrival once; (4) the shelf band with Cancel bottom-left and Remove far right; a LONE repeating selection renders Remove as two lines, 'Skip today' over 'the series continues'. STILL genuinely bulk-only: Break down, Make it tiny, Pin, Steps, Remind me and 'Done on...' never appear here. Bulk actions apply to all and exit select mode.", "Both"),
    ("TOD-13", "Today", "P2", "Move a task (or several) to a chosen day",
     "TWO paths, check both. Single: hold a task, tap 'Move to...' in its own card. Several: hold, 'Select more', tick the rest, then 'Move to...' in the bar. Either way pick 'This weekend', 'Next week', or a calendar day.",
     "One-offs move to that date and wait in the Later list until then (recurring tasks are left alone). REGRESSION CHECK: the SINGLE-task sheet must read 'One task moves to the day you pick.' (never 'move these 0 tasks'), and the move must actually happen. The card (or select mode) closes after.", "Both"),
    ("TOD-20", "Today", "P2", "Combine several tasks into one (the inverse of Break-it-down)",
     "Hold a task, tap 'Select more', tick a second (or more) open one-off, then tap 'Combine'. A cheap Haiku call suggests an umbrella title; edit it (or type your own if the AI is unavailable), then tap 'Combine'.",
     "The 'Combine' action appears only at two or more eligible tasks. The review shows the chosen tasks and an editable name. On accept the originals fold into ONE new task placed at the earliest of their due dates (an undated or due-today task lands it on Today with no imposed deadline), the originals are tucked away (recorded on the umbrella, so nothing is lost), Today declutters from several rows to one, and select mode exits. The umbrella completes like an ordinary task. Needs the Worker deployed with /combine for the AI title; without it you can still type the name.", "Both"),
    ("TOD-20b", "Today", "P2", "Combine WITHOUT AI (name it yourself)",
     "With AI off, hold a task, tap 'Select more', tick a second eligible task, then tap 'Combine'.",
     "'Combine' appears with AI off just as with AI on. The review opens with the name PREFILLED from the selected titles joined by commas (no AI call, no name-finding step), editable. On accept the originals fold into one umbrella exactly as the AI path does (earliest due date, originals recorded for un-combine, select mode exits). Only the source of the suggested name differs.", "Both"),
    ("TOD-21", "Today", "P3", "Combine handles broken-down (decomposed) child tasks",
     "Break a task down into steps, then hold one step, tap 'Select more', tick another, and Combine them.",
     "The steps fold into the umbrella cleanly. If you combine ALL of a decomposed parent's remaining steps, that parent is tidied away too (its work moved into the umbrella). If some steps remain, the parent stays and still completes normally once those are done. Nothing is orphaned or double-counted.", "Both"),
    ("TOD-14", "Today", "P2", "Low-capacity day (gentle recalibration)",
     "Under the weight gauge, tap 'Low on energy? Make it a low day'.",
     "The gauge recalibrates to a gentler capacity (the same task count reads as fuller) and the label gives permission ('A low day. A couple of things is plenty.', up to 'Just pick one, the rest waits.'). A brief affirmation shows. The backlog is untouched, nothing is deferred or shamed. The toggle reads 'Back to a normal day' to undo, and the state self-clears at midnight (per-day, never a setting).", "Both"),
    ("TOD-15", "Today", "P3", "Evening wind-down nudge",
     "Open the app in the evening (after 6pm) with the day not yet closed.",
     "A calm line appears above 'Close the day' ('Evening's here. Close the day when you're ready, even a little counts.'), inviting the closing ritual. It is in-app only (no notification), never shaming, and absent during the day.", "Both"),
    ("TOD-16", "Today", "P2", "Rooms pill, phase greeting, soft cards (Today reborn)",
     "Look at the Today header and the line under 'Today'. Tap the 'Rooms' pill, then a room. Open the app at different times of day.",
     "The header shows the date plus one 'Rooms' pill (three dots and a label), never the old four-link row that wrapped on narrow phones. Tapping Rooms opens a calm bottom sheet listing Repeating, Routines, Calendar, Settings (each with a one-line hint); tap one to go, tap the scrim to close. The greeting under 'Today' changes with the clock: 'Good morning/afternoon. Just today.', 'Winding down. Just today.' in the evening, a restful line late at night. Task rows sit on a soft shadow, floating a hair above the living background.", "Both"),
    ("TOD-17", "Today", "P2", "Pull a Later task forward to today (hold + Move to)",
     "Schedule a task for a future day (capture with 'Date...', or push one to Tomorrow) so it sits in the 'Later' list. Tap-and-hold it, then tap 'Move to...' in its own card and pick 'Today'.",
     "A Later task holds into the SAME in-place card as a Today row. Its decided action set is: Move to... / Break down / Steps / Make it tiny / Mark as a lot / Select more / Remove / Close. Deliberately absent: Pin (Today-only by design, see PIN-05), Tomorrow (an ambiguous pull backward on a future task) and Remind me (the presets are today-shaped). 'Move to...' offers Today alongside This weekend / Next week / a calendar day. Picking Today moves it into Today (its due becomes today) and out of Later. No shame framing, the mirror of pushing to tomorrow.", "Both"),
    ("TOD-18", "Today", "P3", "Secondary actions read as tappable (not labels)",
     "Look at the quiet text actions: 'Done adding' (open the capture drawer), 'Sync across devices' and the daily-reminder line (Today footer), 'Select all' (in select mode), and the low-day toggle.",
     "Each is underlined, so it clearly reads as a tappable link rather than an inert label, while staying calm (soft ink, no mauve). Primary actions stay buttons. Plain labels (the rotating ethos, dates) are never underlined.", "Both"),
    ("TOD-19", "Today", "P2", "Long titles wrap and stay fully visible (incl. with a reminder)",
     "Add a task with a very long title. Set a reminder on it (tap-and-hold, Remind me). Look at the row, at the held card, and at the same task in select mode (reached via 'Select more'). Try it on web and on Android.",
     "The long title wraps onto up to three lines on the calm row and stays fully visible, including when a reminder bell shares the row (it never collapses to a blank line). It behaves identically with or without the reminder, in select mode, and on both platforms. No scrolling, just a calm static wrap. The HELD CARD clamps the same title at two lines (intended: the card's job is the actions, and the full title is one tap away on the row behind it).", "Both"),
    ("VIS-04", "Visual", "P2", "The living background (Today's signature, calm, reduced-motion aware)",
     "Open Today at different times of day, on light and dark. Turn on Reduce Motion (in Settings or the OS) and watch the Today background. Then move to Routines, Calendar, Settings.",
     "On Today, a soft time-of-day gradient (dawn / day / dusk / night) with a warm top glow and a softer lower pool sits behind the screen. It only ever shows in the margins: cards and rows stay on near-opaque surfaces, so text is always full-contrast, never washed out. With Reduce Motion on, the colour still resolves to the time of day but the drift stops. The other screens (Routines, Calendar, Settings, etc.) sit on a solid, calm Dusk background, with no grey flash.", "Both"),

    # --- OCD reassurance (Cluster A) -----------------------------------------
    ("OCD-01", "Today", "P2", "Done is done (the rotating completion affirmation)",
     "Fully restart the app, then complete SEVERAL tasks in a row and read the line under each: tap a Today task, and also hold a task -> 'Select more' -> tick some -> Done. Keep going past eight completions to see the pool wrap. Then restart the app and complete one more.",
     "A brief calm line appears near the capture and clears itself after ~3.5 seconds. It ROTATES IN ORDER through a small fixed pool of eight ('Done is done. Recorded.', 'Filed. You can stop checking it now.', 'Good enough is done. Let it go.', and five more), wrapping back to the first after the eighth. This is DETERMINISTIC, not random: the same sequence every time, and because the counter resets on restart the first line after opening the app is always 'Done is done. Recorded.'. That predictability is the point. Rotation gives variety of WORDING, never a variable reward: assert there is no randomness, no score, no count, no streak, no escalating praise, no exclamation mark, and that every line sits in the same calm register so there is nothing to chase. Reduce-motion safe, never shaming. NOTE a whole-task finish shows the bloom INSTEAD of this line, never both (see the celebration cases).", "Both"),
    ("OCD-02", "Today", "P2", "Good enough (the release lives in the affirmation, not a button)",
     "Complete several tasks in a row on Today and read the quiet line under each. There is deliberately NO 'Good enough' action on the held card or anywhere else: hold a task and confirm none is offered.",
     "The perfectionism release is carried by the rotating completion affirmations ('Good enough is done. Let it go.', 'Filed. You can stop checking.'), which is permission to release without adding a twelfth control to the held card. Holding a task offers no 'Good enough' anywhere; that is correct, not a defect.", "Both"),

    # --- Routines (Cluster D) ------------------------------------------------
    ("RTN-01", "Routines", "P2", "Create and run a routine",
     "From the Today header tap Routines, then + New routine. Name it (e.g. Morning), pick Morning / Evening / Anytime, type a few steps one per line, tap Add routine. Then tick some steps.",
     "The routine appears grouped under its time-of-day with an 'N of M' progress. Tapping a step marks it done for today (a sage tick and a strike-through) and updates the count. Calm, with no streak and no celebration pressure.", "Both"),
    ("RTN-02", "Routines", "P2", "A routine is fresh tomorrow (never a streak)",
     "Tick some routine steps today, then advance the device clock to tomorrow and reopen Routines.",
     "Every step is un-ticked again and the progress is back to 0 of M. There is NO streak count, no 'you missed it', and no chain to break: yesterday simply falls away with no guilt (the never-shame spine).", "Both"),
    ("RTN-03", "Routines", "P3", "Remove a routine (recoverable)",
     "On a routine card tap Remove, then optionally tap Undo.",
     "The routine is removed with a brief 'Routine removed. Undo' banner, not a confirmation dialog. Tapping Undo within a few seconds restores it, otherwise it stays gone. Recoverable, never a confirm gauntlet.", "Both"),
    ("RTN-04", "Routines", "P2", "Edit a routine (today's ticks survive)",
     "On a routine card tick one step, then tap Edit. Change the name, add a step on a new line, remove another, and tap 'Save changes'.",
     "The form opens prefilled (name, when, steps one per line, nudge hour). After saving, kept steps keep today's ticks (the ticked step stays ticked), the added step appears unticked, and the removed step's tick never resurrects. Supports building a routine slowly, start with 3 and grow it.", "Both"),
    ("RTN-05", "Routines", "P2", "A daily nudge per routine (opt-in)",
     "Create or edit a routine, turn the nudge on, pick an hour (e.g. 8:00 pm), save. On Android, wait for the hour (or check the scheduled notification).",
     "Off by default. When on, ONE gentle daily notification fires around the chosen hour: title = the routine's name, body 'When you're ready. A step is plenty.' Same calm channel as the daily reminder, inexact timing is fine. Removing the routine (or turning the nudge off) cancels it. On web the nudge row explains reminders aren't available on this device, and the routine still saves. If notification permission is denied, a calm reason line shows, never a dialog.", "Android"),
    ("RTN-06", "Routines", "P2", "Save requirements are visible, never silent",
     "Open the routine form and tap save with no name; then with a name but no steps.",
     "The app answers every tap: a missing name focuses the name field with 'Give it a name first, anything works.'; missing steps focus the steps box with 'Add at least one step, one line is plenty.' Each hint clears on the first keystroke. No disabled buttons, no dialogs, never a red shout.", "Both"),
    ("RTN-07", "Routines", "P2", "Minute-level nudge time, honestly 'around'",
     "In the routine form turn the nudge on, tap the time, and enter 20:47 in the hour/minute boxes (a 24-hour hint is shown). Save. Check the routine card.",
     "The live line and the card read 'Nudge around 8:47 pm' (or 20:47 per the device's convention), the word 'around' is deliberate: Android delivers inexactly and the copy never promises the minute. On Android the notification arrives near the chosen minute; on web the calm 'not available on this device' line shows and the routine still saves.", "Android"),
    # --- Rhythms (nudge-only recurring self-care) --------------------------
    ("RHY-01", "Rhythms", "P1", "Add a Rhythm from a preset, nudge-only, never a task",
     "On the Routines screen scroll to Rhythms, tap 'Water, around every 2 hours'. On Android, with the app closed, wait for a window hour.",
     "A calm card appears in the Rhythms section with a cadence line ('Around every 2 hours, 9am to 9pm') and NO checkboxes, NO 'N of M' progress, and NO streak, ever. It never appears on the Today list. On Android a gentle water nudge arrives around a window hour on the calm (silent, no-badge) channel, and after firing it leaves no entry anywhere in the app (nothing to feel behind on). On web the section says reminders arrive on the phone, and it still saves.", "Both"),
    ("RHY-02", "Rhythms", "P2", "Make a custom Rhythm, and edit a preset (fully flexible, minutes-granular)",
     "Tap '+ New rhythm'. Name it (e.g. Stretch), walk 'How often' with the stepper across its full range (30 min at one end, 12 h at the other; the ladder is 30/45 min, 1 h, 1 h 30, 2 h, 2 h 30, 3, 4, 5, 6, 8, 12 h) and the 'Active hours' window (start / end), watch the live cadence preview, tap 'Add rhythm'. Then tap Edit on the Water card, change its interval, window or name, and Save changes.",
     "The stepper walks the whole ladder with plain labels ('every 30 minutes', 'every hour', 'every 1 h 30 min', 'every 2 hours') and stops calmly at both ends (the minus greys out at 30 minutes, the plus at 12 hours). The live preview matches ('Around every 1 h 30 min, 9 am to 9 pm'), and the saved card shows the same line. A 90-minute Rhythm actually fires off the hour (9:00, 10:30, 12:00...). A missing name focuses the name field with a calm hint, never a silent bounce. Editing a preset opens the same form prefilled (a Rhythm made before the minutes ladder opens at its same cadence, nothing shifts) and updates it in place. Any change reschedules immediately AND the 'Next nudge around {time}.' line under Rhythms updates in place right after the save / pause / remove / undo, no screen re-entry needed. Still no checkboxes, no progress, no streak.", "Both"),
    ("RHY-03", "Rhythms", "P2", "Pause, resume, and remove a Rhythm (recoverable)",
     "On a Rhythm card tap Pause, then Resume. Then tap Remove, and Undo within ~6 seconds.",
     "Pause shows a quiet 'Paused' on the card and (on Android) stops the nudges with no fuss; Resume brings them back. It is an indefinite, honest pause, never a timed snooze that could silently die. Remove takes the card with a brief 'Routine removed. Undo' banner (not a dialog); Undo restores the Rhythm and its nudges. Nothing is ever shamed.", "Both"),
    ("RHY-05", "Rhythms", "P1", "Rhythm nudges actually arrive, visibly, and the health line tells the truth",
     "On Android with at least one active Rhythm: check the faint line under the Rhythms section ('Next nudge around {time}.'). Tap 'Allow alarms & reminders' and flip the toggle ON (Android 14+ ships it off; on 12/13 it should already be on). Return to the app, force-close it, wait for the next slot, and confirm a nudge PEEKS (heads-up banner, not just a silent tray item) WITHOUT opening the app. Then reopen and re-check the line.",
     "The health line is ONE calm sentence, 'Next nudge around {time}.', with the right soonest time; deliberately NO scheduled-notification count (an honest '39' reads as overwhelm for exactly this audience). The one red-flag state remains: when NOTHING is scheduled while Rhythms exist, the line says so plainly and suggests re-saving a rhythm. With 'Alarms & reminders' allowed, nudges are EXACT alarms: they arrive on time with the app fully closed, through Doze (this was the launch-week 'nudges only fire when I open the app' bug: without the grant, every alarm is inexact and the OS holds it until the app next wakes, battery settings notwithstanding). The 'Allow alarms & reminders' link opens DoubleDone's own toggle (falling back to the all-apps list on OEMs that block the direct screen). Nudges peek on the HIGH-importance 'Rhythms' channel. Coming BACK from the toggle re-arms every nudge immediately (no app restart needed), and reopening the app re-schedules everything quietly anyway (the resilience sweep), which NEVER shows a permission prompt. The battery hint remains the second lever for aggressive OEMs.", "Android"),
    ("RHY-04", "Rhythms", "P1", "Fixed-time Rhythm (meds): nudges at set clock times",
     "Tap the 'Meds, at 8 am and 8 pm' preset. Then Edit it: in the form switch between 'Every so often' and 'At set times', add a time (24-hour entry, e.g. 14:30), remove one, and save. On Android, wait for a set time to pass.",
     "One tap creates a Meds Rhythm firing at 8:00 and 20:00 (the card reads 'At 8:00 am · 8:00 pm' in the device's time convention). The form is EITHER every-N-hours OR at-set-times, one mode at a time; times are 24-hour entry with a live preview that sorts them into clock order, duplicates collapse, at most 8 times, and saving with no valid time shows a calm 'Add at least one time.' hint, never a silent nothing. On Android each set time delivers one calm nudge (title = the Rhythm's name, DEFAULT importance, delivered 'around' the time, inexactly). Still nudge-only: never a task, never ticked, no streak anywhere, and Pause / Resume / Remove behave exactly as RHY-03.", "Both"),
    ("SB-09", "Scrapbook", "P2", "Keepsakes follow the account across devices",
     "Signed in on two devices (e.g. Android + web). Make a keepsake on device 1, then open the app on device 2 (Today first, then the Calendar). Remake the same week's keepsake on device 2 and check device 1 after its next open. Prerequisite: an account signed in on both.",
     "The keepsake appears on device 2's Calendar for its week (sync runs on app open; the Calendar shows it on next visit). A remade week's newer keepsake replaces the older one everywhere, never duplicates. Legacy device-local keepsakes (made before cloud persistence, stored as raw image data) deliberately do NOT sync and stay only where they were made. Signing into a DIFFERENT account never migrates the previous account's keepsakes up (the cross-account guard). A scrapbook sync failure is silent and never breaks task sync or shows an error.", "Both"),
    ("REP-01", "Repeating", "P1", "Removing a repeating task from Today skips ONLY today",
     "Long-press a due repeating task on Today, tap Remove (single, in place; and in bulk via 'Select more').",
     "Today manages days: the instance vanishes from Today with 'Skipped just for today. The repeat continues.', the series survives (still in the Repeating drawer), and it returns on its next due day. A mixed selection tombstones the one-offs and skips the recurring. Screen readers hear 'Remove today's instance of {title}' on BOTH the held card and the bulk bar. The recurring card also shows NO Tomorrow and NO Move to... (re-dating silently does nothing to a repeat, and an action that does nothing is worse than an absent one), but DOES offer 'Mark as a lot' (a chore can absolutely be a lot).", "Both"),
    ("REP-02", "Repeating", "P2", "Edit a series from the Repeating drawer",
     "Open the Repeating drawer, tap Edit on a series. Change the title, the cadence (daily / weekdays / every N), and the start date. Save.",
     "The drawer manages the series: the edit sheet prefills everything, saving updates the series in place (shape identical to a captured one), and Today reflects the new cadence. An empty title cannot save.", "Both"),
    ("REP-03", "Repeating", "P2", "Remove a series from the drawer (recoverable)",
     "In the Repeating drawer tap Remove on a series, then Undo within ~6 seconds.",
     "The series tombstones with a calm 'Repeating task removed. Undo' bar, never a confirm dialog. Undo restores it fully (and re-syncs). Letting the bar lapse keeps it removed.", "Both"),
    ("REP-04", "Repeating", "P2", "The drawer says plainly that a repeating task does not notify you",
     "Open Menu, then Repeating, and read the two lines under the heading. Follow what the second one tells you to do: hold a task on Today and look for 'Remind me', and open Settings and look for the daily nudge.",
     "Under the subtitle sits a fainter clarification: 'A repeating task APPEARS on Today when it's due. It doesn't notify you. If you want to be told, hold the task on Today and choose Remind me, or turn on the daily nudge in Settings.' Both routes it names really exist and really do notify. It reads as a clarification, not a warning, so it is fainter than the subtitle and nothing about it looks like an error. The reason it exists: the app calls a repeating TASK and a NOTIFICATION both a kind of 'reminder', so someone who sets a task to repeat can reasonably expect to be told when it is due, and then is not. That is our confusion to fix, not theirs to work out, and it is fixed by stating the limit where the expectation is formed rather than by renaming anything ('Repeating' is the right word).", "Both"),
    ("LB-10", "Calendar", "P2", "Repeating tasks project onto the calendar ('hair washing day')",
     "Create a repeating task with an interval (e.g. 'Wash hair' every 4 days) and a daily one. Open the Calendar and look at the days ahead. Tap a future day that carries the planned dot. Then skip-today one occurrence (hold the task on Today on its due day, Remove) and re-check that day in the calendar.",
     "Every future day a repeat lands on carries the planned (outline) dot, exactly like a dated one-off, so 'is it hair washing day?' is answered by a glance at the calendar (the power user's ask, 2026-07). Tapping the day lists the repeat under the planned heading with the ○ mark and a small ↻ so a repeat reads as a repeat. A skip-today'd occurrence does NOT mark its day (the series continues; that day is not planned). Today and past days never carry planned marks: the past belongs to completions, today to Today. One-off planned dots and the day-detail behave exactly as before.", "Both"),
    ("LB-11", "Calendar", "P1", "Tap a future day to add something to it",
     "Open the Calendar. Tap a day AHEAD and read what the day offers. Add something for it, then check Today's Later list (or re-open that day). Then tap a PAST day and today ITSELF and look for the same offer. Finally, open the add input on one future day, type a few words WITHOUT adding, then tap a different day and check what happened to what you typed.",
     "A FUTURE day offers 'Add for this day' (plain accent text, never a filled button; this is an offer on a screen for reflection, not the loudest thing on it). Tapping it opens ONE line with the placeholder 'One thing for this day', plus Cancel and Add; there is deliberately NO date picker, because the day you tapped IS the date, which is the whole point of arriving here by tapping it. Add is disabled until something is typed. Adding drops an ordinary one-off task dated to that day: it appears immediately under that day's SCHEDULED list with the ○ mark, it shows in Today's Later view, it syncs and exports like any other task, and a quiet 'Added for <day>.' confirms it landed on the day you chose (you are still looking at the calendar, not at Today, so the confirmation names the date). The offer returns straight away so a second thing can be added. A PAST day offers NOTHING: the past is a shame-free RECORD, and back-filling it would either lie about when the task was made or invite padding a day to look busier. TODAY offers nothing either, and that is deliberate rather than an oversight: Today already has capture permanently docked, and a second door to the same action in a different place teaches two habits for one thing. Switching days mid-type DISCARDS what was typed and clears any confirmation, because the input belongs to the day it was opened on and silently re-pointing it at a new date would land a task somewhere nobody asked for. This closes the promise the 'Lookback -> Calendar' rename made: the name says you can schedule into a day, and until now the forward view was read-only.", "Both"),

    ("TOD-25", "Today", "P2", "The closed day shows a calm forward COUNT, never the list",
     "With tasks due tomorrow and in a few days, close the day. Read the rested screen. Tap the waiting-ahead line. Then reopen the day, remove all future tasks, close again.",
     "The rested screen shows one quiet underlined line: 'N things are waiting for the days ahead. They are safe.' (singular at 1). It is a COUNT only: no future task titles appear anywhere on the closed screen, because closing the day means setting it down, and the line exists to answer 'did I actually add that?' without re-loading tomorrow's weight. Tapping it opens the Calendar. With nothing upcoming, the line is absent entirely. (Found 2026-07-24: closing the day used to hide ALL future tasks with no reassurance at all.)", "Both"),

    ("TOD-24", "Today", "P2", "Edit a task's title from the held card",
     "Hold a task. On its card, the title carries a faint underline: tap it, type a new name, press enter (or tap away). Also try: clearing the text entirely and tapping away; holding a SLICED task (the title shows '· n / N') and editing it; and editing a repeating task.",
     "Tapping the title turns it into a text field in place, pre-filled with the RAW title (never the '· n / N' step counter, which is display-only). Enter or tapping away saves; the row updates immediately and the card STAYS open (fix a typo, then keep acting). An emptied or unchanged title saves nothing (no sync write). Renaming a repeating task renames the series, as the single visible row implies. The rename syncs across devices like any edit. Screen readers hear 'Edit the title: {title}' on the tap target.", "Both"),

    ("TOD-26", "Today", "P2", "Held-card action labels render in full, never a clipped glyph",
     "On a DENSE Android device (Samsung S22, 3x density, is the known repro) hold a task and read EVERY row of the card: the lead actions and their sub-labels, and the terminal row 'Close / Select more / Remove'. Repeat with the system font size bumped up, and in a longer locale (fr, it).",
     "Every label renders whole, with no trailing glyph clipped ('Remove' never reads 'Remov'). The design-1a card is a stack of full-width rows (label left, sub-label/state right), and the terminal row is space-between (Close far left on the accent, Select more in the middle, Remove muted-brick far right), so no label's width is ever computed from fractional leftover space (the bug the old equal-width grid + flex-spacer had, reported on an S22). Nothing overflows or wraps oddly at large font.", "Android"),

    ("TOD-23", "Today", "P2", "Done on… corrects a COMPLETED task's day",
     "Tap a rolled-over task done (it records today). Long-press the completed task: its card is deliberately SMALL, offering only 'Done on…' then Select more / Remove / Close, with no Done at all (tapping the row is the one way to finish a thing) and no shaping actions (a finished thing needs nothing shaped). Pick a past day (yesterday to 14 back; today is not offered). Also long-press an OPEN task and confirm 'Done on…' is NOT offered there.",
     "Correction is a property of a completed task: the natural flow is tap-done first, refine after. Picking a day re-stamps the completion to that day's noon; the Calendar attributes it honestly on every device; the affirm reads 'Recorded for {day}. Your Calendar tells it true.', no bloom, no haptic. Recurring tasks never offer it.", "Both"),

    # --- Haptics (Android device only) ---------------------------------------
    ("HAP-01", "Haptics", "P3", "Earned-moment haptics fire (Android)",
     "On a physical Android device with a haptic motor: complete a single task; clear the whole day; close the day with Goodnight; break a dreaded task into steps; and (premium) reveal a scrapbook.",
     "Soft tap on a single completion; a fuller success buzz when the day clears; a warm soft tap on Goodnight; a light tap when steps land; a success flourish when the scrapbook image appears. Nothing buzzes on plain taps, navigation, capture, or any error.", "Android"),
    ("HAP-02", "Haptics", "P2", "Reduced motion silences haptics",
     "Set Settings -> Motion -> Reduce (or enable the OS reduce-motion), then complete a task and close the day.",
     "No haptic fires for any cue while motion is reduced (the accessibility guarantee). The Motion hint states Reduce also stops the buzz.", "Android"),

    # --- Android native polish (device only) ---------------------------------
    ("AND-01", "Android", "P3", "Screen stays awake in Focus mode",
     "On an Android device, open Focus (Focus on one thing) and leave the screen untouched past the usual sleep timeout.",
     "The screen stays on while Focus is open, and returns to normal sleep behaviour once Focus is closed.", "Android"),
    ("AND-02", "Android", "P3", "System bars match the theme",
     "On Android, switch the in-app theme (Settings) between light and dark, including a case where the app theme differs from the system theme. Watch the status bar and the bottom navigation bar.",
     "Status-bar and navigation-bar icons stay legible against the app background in both themes; no white flash on launch or overscroll.", "Android"),
    ("AND-03", "Android", "P2", "Launcher long-press shortcuts",
     "On Android, long-press the DoubleDone home-screen icon. Tap 'Brain dump'; relaunch and tap 'Focus on one thing'. Try both from a cold start and with the app backgrounded.",
     "'Brain dump' opens the app with the capture box focused and ready to type. 'Focus on one thing' opens directly in Focus mode.", "Android"),
    ("AND-04", "Android", "P2", "Share text into DoubleDone (Android sheet + installed web app)",
     "On Android, from another app (a browser, notes, a chat) use the system Share sheet and pick DoubleDone. Try sharing a line of text, and a URL. Then on the WEB: install doubledone.app (browser menu -> Install app / Add to Home Screen), share a page or text from another app, and pick the installed DoubleDone.",
     "Both paths behave identically: DoubleDone opens with the shared text already in the capture box on Today, OPEN, focused and ready to add, including when the capture box was collapsed AND on a cold launch (regression: the parked seed was cleared before the box mounted, so a share only worked if the box was already open). The share is cleaned to ONE calm line: a browser's quoted-selection share ('\"Title\"\\n https://site#:~:text=...') lands as just the title, links are dropped when words exist, a bare link keeps the link with the highlight fragment stripped. Adding it makes a normal task; nothing is sent anywhere until you do. The web path needs the PWA installed (that is how browsers expose share targets).", "Both"),
    ("AND-07", "Android", "P2", "Missed nudges never pile into a guilt-heap",
     "On a device with an active Rhythm (e.g. water every 30 min) and the daily reminder on, leave the phone alone until several nudges have fired unnoticed and sit in the notification tray. Then open DoubleDone. Also set a per-TASK reminder, let it fire, and open the app.",
     "On opening the app, every delivered Rhythm and daily/routine notification is dismissed from the tray automatically: they are offers to open the app, and the app is open, so a pile of missed ones must never sit there reading as a guilt-heap (the never-shame rule; reported in the wild 2026-07-24). The per-TASK reminder is deliberately KEPT in the tray, it points at one specific task and stays actionable. Works on iOS too (matched by identifier there, since iOS has no channels). What this does NOT yet do: expire nudges while the app stays closed (needs a native config plugin; see the Backlog).", "Both"),

    ("AND-05", "Android", "P1", "Home-screen widget renders Today (re-enabled + self-diagnosing)",
     "On Android, long-press the home screen, open the widget picker, and place the DoubleDone 'Today' widget. Read what it draws, and look at its SHAPE (all four corners, and how much empty space sits under the last task). Drag its resize handles taller and wider. Then add or complete a task in the app and re-check the widget after a moment (it also self-updates every ~30 min).",
     "The widget appears in the picker (label 'DoubleDone: Today') and, once placed, draws a card: the 'Today' header plus up to four unfinished task titles, or a calm rested line ('All done for today.', 'Closed for today.', 'Nothing for today yet.'). Tapping anywhere opens the app. The widget was re-enabled 2026-07-25 after the original blank render was traced to an unproven library theory, not a confirmed cause; the headless task is now SELF-DIAGNOSING, so instead of a silent blank a failure draws its own error text on the widget ('DoubleDone' + the error). Decision tree if it misbehaves: tasks shown = working; an error string shown = we have the exact cause on screen (no adb needed); still fully blank = the headless task is not firing at all (a registration issue). SHAPE: the card is rounded on ALL FOUR corners (never rounded-on-top-but-sliced-flat-across-the-bottom, the 'tombstone' the match_parent card used to draw), and it hugs its content, so three tasks make a SHORT card, never a tall one with a big empty field underneath. Resizing is the interesting one: dragging the widget TALLER shows MORE TASKS (the line count follows the height), and dragging it shorter shows fewer, down to a minimum of one. THE CARD IS NEVER CLIPPED, at any size, any task count, any system font size: check the bottom edge at every step of a drag and after adding tasks one at a time up to 20. A correct card ends with a fully-drawn last line and four symmetric rounded corners; a FAILED one is rounded on top, flat across the bottom, with the last title sliced through the middle (the 'tombstone'). Regression, device-reported TWICE (2026-07-25): the line budget was computed assuming the system font size was at its default, so on a phone set larger the content grew past the slot and Android sliced the rounded bottom off. The budget now scales with the device's font setting and always leaves half a line of air unspent, so it errs toward showing ONE FEWER task rather than overflowing. That trade is deliberate: an under-filled card is invisible, an over-filled one is mutilated. Because of it, a tall widget may show slightly fewer tasks than would technically fit, which is correct and not a bug. Anything hidden is always accounted for by the '+n more' line, so the count is honest even when the list is trimmed. It never grows dead space, because the card is content-height, and it never spills past its slot, because when tasks overflow the budget the last row becomes '+n more' rather than a fifth title. Widen it and titles simply get more room before truncating. VERIFIED WORKING on device 2026-07-25 (v15); the original blank was the React Compiler injecting a hook into TodayWidget, fixed with 'use no memo', NOT the new-architecture theory it was blamed on for a month.", "Android"),
    ("AND-08", "Android", "P2", "The always-light widget, for a dark wallpaper",
     "Set a DARK wallpaper and put the phone in dark mode (the combination the default cannot serve). Open the widget picker: TWO DoubleDone widgets are offered. Place 'DoubleDone: Today (light)' and compare it against 'DoubleDone: Today'. Then flip the phone to light mode and re-check both. Complete a task in the app and confirm BOTH placed widgets refresh.",
     "The picker offers 'DoubleDone: Today' (follows the phone's light/dark setting) and 'DoubleDone: Today (light)'. The light variant draws the warm paper card with dark ink ALWAYS, in either system mode, so it stays legible over a dark wallpaper where the default's dark card would sink into the background. The default still follows the system as before. Both carry the same content, the same 'Today' + DoubleDone header, and both refresh together when tasks change (the updater fans out over every registered widget name). Placing both at once is fine.", "Android"),
    ("AND-10", "Android", "P1", "A missed nudge expires from the tray by itself, app closed",
     "With the app CLOSED (swiped away, not just backgrounded), let a daily reminder or a Rhythm nudge arrive and DO NOT touch it or the app. For a Rhythm with a short interval (e.g. every 30 minutes), wait past the next slot. For the daily reminder, check the tray after ~12 hours. Also set a per-task 'Remind me' nudge, let it arrive, and leave it overnight.",
     "A swept-class nudge (the daily reminder, a routine nudge, a Rhythm slot) removes ITSELF from the tray without the app ever opening: a Rhythm slot disappears no later than the moment its next slot arrives (so two of the same Rhythm are never visible at once), and any of them is gone within 12 hours at most. This is the closed-app half of 'missed nudges never stack into guilt'; the app-open sweep (AND: open the app, tray clears) remains the other half and still works. The PER-TASK nudge deliberately does NOT expire: it is actionable, not a come-back invitation, and it stays until acted on or swept never (same rule as the app-open sweep). Mechanism, for debugging: the JS payload carries timeoutAfterMs and a patch-package patch on expo-notifications hands it to Android setTimeoutAfter, so if expiry stops working after an expo-notifications upgrade, check that patches/expo-notifications*.patch still applies (a failed patch FAILS the install loudly, it cannot silently vanish). Requires Android 8+ (API 26); older Androids simply keep the old behaviour.", "Android"),
    ("AND-09", "Android", "P1", "The app offers its own lifelines, once each, never two at a time",
     "On a device with NO DoubleDone widget on the home screen and the daily reminder OFF, close the day and read the rested screen. Answer (or dismiss) whatever it offers, then close the day again on a later day and read it again. Then place a widget, clear the offer flags (reinstall or clear data), and close the day once more. Also check the rested screen on iOS and on the web app.",
     "The rested screen offers at most ONE lifeline, ever: first the daily reminder ('Want one gentle nudge a day to come back?'), and only on a LATER evening, once that ask is spent, the widget ('Want today on your home screen?' -> 'Show me how' reveals 'Hold your home screen, tap Widgets, then find DoubleDone'). Two asks NEVER appear together, because the goodnight screen is for setting the day down, not for being sold to. Each is shown once and never returns whatever the answer was. Someone who ALREADY has a widget on their home screen is never asked (the app checks the launcher), and the widget offer never appears on iOS or web, where there is no widget to place. This is the 'app comes to you' pass: a returning user asked for four things that already existed because every lifeline is opt-in and none was ever surfaced.", "Android"),
    ("AND-06", "Android", "P2", "Remind me in X hours (per-task nudge)",
     "On Android, tap-and-hold a today task, tap 'Remind me', pick a preset (e.g. 'In 1 hour'). Lock the screen and wait. Then set another and let its time pass WITHOUT completing the task. Separately, complete / remove / push-to-tomorrow a task with a pending nudge, and open 'Remind me' after 9pm. NOTE: if nothing fires, confirm Settings -> Apps -> DoubleDone -> Battery is not 'Restricted' (Samsung One UI throttles alarms).",
     "The notification reliably FIRES at the chosen time as a heads-up (the task as the title, 'Whenever you are ready.' as the body), even with the screen off, via an exact alarm. The row shows a small bell + time. Once the time has passed the bell clears on its own (on the next open or app resume) even if the task is still open. Completing / removing / deferring cancels a pending nudge (no poke about a handled task). After 9pm the late presets are hidden. Web does not show 'Remind me'.", "Android"),

    # --- Web push (deployed; needs VAPID configured) -------------------------
    ("WEB-01", "Web", "P2", "Daily reminder via web push",
     "On the deployed web app (PC or Android Chrome) with VAPID configured, turn the daily reminder on (the 'Turn on daily reminder' action in the Today footer, or Settings > Daily reminder) and allow notifications. Check around your daily hour (the hourly cron can be run manually to verify without waiting).",
     "Toggling on registers a service worker and subscribes the browser; a calm 'Your today is here when you are ready.' notification arrives around the daily hour, and tapping it opens the app. Toggling off unsubscribes. The push carries no task content. The toggle is hidden when VAPID is unconfigured.", "Web"),

    # --- Web landing page (the front door) -----------------------------------
    ("LP-01", "Landing", "P1", "The web landing is the front door",
     "On the web in a fresh browser (or after clearing 'doubledone.onboarded.v1'), open doubledone.app/. Read the page, then tap Begin. Separately, as a returning (onboarded) user, open doubledone.app/. On Android, just open the app.",
     "A calm marketing page renders at the root: the wordmark, 'Today is finite and achievable.', the never-shame promise, the audience line (ADHD / autism / OCD), the three-step loop, the Calendar payoff, and a Begin CTA. Begin opens the app at /today (a first-timer continues into the welcome). A returning, onboarded visitor is redirected straight to /today and never sees the landing. On native the app opens to Today, never the landing. Checkout still returns to /premium and deep links still resolve.", "Both"),

    # --- Onboarding: the one-time guided welcome ------------------------------
    ("ONB-01", "Onboarding", "P1", "Guided welcome on first run (7-screen)",
     "On a fresh install (or after clearing 'doubledone.onboarded.v1'), open the app. Walk Begin -> type a few lines -> Sort it for me -> This looks right -> Got it -> Almost there -> a Premium teaser -> Open Today. Try Back to a previous screen, and separately try Skip.",
     "Today redirects to the welcome exactly once: a 7-screen sequence with a quiet 7-dot progress, Skip on every screen but the last, and Back (top-left) from screen 2 on with the typed text intact. The dump is triaged in the AI's suggested order but EVERYTHING lands on Today, nothing is pushed to Later on first contact (before trust exists, a vanished line reads as data loss); the reveal teaches 'Everything starts on today. Move anything to tomorrow, or later, whenever you like.'; any big one is flagged 'Looks big, break it down?'; the safety-net and 'what you keep' screens follow as calm info; the penultimate screen is a calm one-screen teaser of what Premium adds (never a hard gate, fully skippable); Open Today saves the tasks and opens Today. Skip, or an empty Sort, leaves immediately, saving whatever was revealed. Never reappears once done. If the AI is slow or offline, everything lands on Today, nothing lost.", "Both"),
    ("ONB-02", "Onboarding", "P2", "Replay the welcome from Settings (non-destructive)",
     "With tasks already on Today, open Settings -> 'See the welcome again'. Walk Begin -> dump a couple of lines -> Sort it for me -> This looks right -> through to Open Today.",
     "The welcome replays identically (all 7 screens), but the new tasks MERGE into the existing list (nothing overwritten) and the onboarded flag is untouched. Skipping returns to Today with no change.", "Both"),
    ("ONB-03", "Onboarding", "P1", "Choose AI-free in the introduction (zero egress)",
     "On a fresh run, Begin, then on the capture screen type a couple of lines. Read the line under the box, then tap 'I'll sort it myself' instead of 'Sort for me'. Walk through to Open Today. To prove zero egress, watch the network: nothing should reach the AI backend.",
     "The capture screen always shows one neutral line naming what each button does ('Sort for me sends these lines to Claude to order your day. I'll sort it myself keeps everything on this device'). Tapping 'I'll sort it myself' sets AI off, puts every line on today with NO call to the AI backend, and the reveal adds 'Sorted on your device, all on today for now.' The final screen reads 'Private by default, nothing leaves your device' (now literally true). AI stays off afterwards (Settings shows Off). Tapping 'Sort for me' instead gives the normal AI triage. Replaying the welcome while AI is off shows the primary as 'Put them on today' and the link as 'Change in Settings'.", "Both"),
    ("ONB-04", "Onboarding", "P1", "The last screen offers the daily nudge, once, sharing its budget with the evening offer",
     "On a FRESH install walk the introduction to the last screen ('That's it. No setup.') and read what it offers. Tap 'No thanks'. Replay the introduction from Settings and check the last screen again. Then close a day and look at the rested screen. Repeat the whole thing on a second fresh install, this time tapping 'Yes, remind me', and check Settings afterwards. On ANDROID, also read the small print on that last screen. On iOS and web, check the same line is absent.",
     "The last screen offers ONE thing: 'Want one gentle nudge a day to come back? You can change or stop it any time in Settings.' with 'Yes, remind me' and 'No thanks'. It is an aside, quieter than the 'Open Today' button, which stays the loudest thing on the screen. This is deliberately NOT a new step: the heading is 'That's it. No setup.', and an eighth screen asking you to configure something would make the app contradict itself on the way out. EITHER answer spends the offer for good, and it is the SAME one-time budget the rested-screen offer uses, so a person is never asked twice across the two surfaces: after answering here, close-the-day never raises it, and a replay of the introduction never re-asks. 'Yes' turns the daily reminder on (Settings shows it On) and the line becomes 'That's set. One quiet nudge a day.'; if the OS permission is refused, the plain reason is shown and the offer is still spent, because asking again after someone has already said no to the system dialog is nagging. On ANDROID only, a small line also MENTIONS the home-screen widget and how to place it; it is a mention, NOT an offer, and it deliberately does NOT spend the widget offer, because an empty widget on day one sells nothing and the real ask should wait for the rested screen when there is a day worth looking at in it. That line is absent on iOS and web, where there is no widget.", "Both"),

    ("ONB-05", "Onboarding", "P2", "The 'what you keep' screen mentions the free monthly scrapbook (AI on only)",
     "On a fresh run with AI on, walk to the 'What you finish, you keep.' screen and read its lines. Repeat on a fresh run choosing 'I'll sort it myself' (AI off).",
     "With AI on, a fourth calm line says a finished week can become a small keepsake image once a month, free. With AI off the line is absent entirely (the scrapbook is an AI feature and must never be pitched to someone who opted out). No button, no link, no step added: it is one sentence inside the existing payoff screen.", "Both"),

    # --- AI: Bite the Elephant (decompose) -----------------------------------
    ("AI-01", "AI decompose", "P1", "Break down a dreaded task",
     "Capture 'Do my taxes', tap Break it down, answer any clarifying questions.",
     "Returns small atomic steps and drops them into Today.", "Both"),
    ("AI-01b", "AI decompose", "P1", "Break it down WITHOUT AI (type the steps yourself)",
     "With AI off (Settings, AI, Turn AI off), long-press a task and tap 'Break it down'.",
     "Instead of the AI flow, a calm modal opens ('Break it into steps') naming the task. Type one step per line and tap 'Break it down'. Each line becomes a child step, the task goes silent (hidden) and is done only when every step is finished, exactly like an AI breakdown. NO network call is made, and empty input does nothing. The 'Break it down' affordance appears in the same place whether AI is on or off.", "Both"),
    ("AI-02", "AI decompose", "P2", "Time estimate shows",
     "On the breakdown review, look for the pace/time estimate.",
     "A sensible total estimate is shown (the crowd/pace estimate).", "Both"),
    ("AI-03", "AI decompose", "P2", "AI egress disclosure at point of use",
     "Open the Break-it-down questions modal.",
     "A calm one-liner discloses the text is sent to an AI and kept anonymously.", "Both"),
    ("AI-04", "AI decompose", "P2", "Friendly error state",
     "Turn off wifi (or block the AI URL), then Break it down.",
     "A calm friendly error. No raw HTTP/stack. App stays usable.", "Both"),
    ("AI-07", "AI decompose", "P2", "Breakdown keeps the real task as a silent parent (chain)",
     "Break down a task (e.g. 'Plan the party'), then complete all of its steps, in any order.",
     "The original task disappears from Today and Later (it becomes a silent parent, not clutter beside its steps). When the last step is done, the real task completes on its own with the held 'you finished the whole thing' bloom (see AI-09) and lands in the Calendar as the finished real task. Multi-phase: finishing a milestone's steps cascades up to the root.", "Both"),
    ("AI-08", "AI decompose", "P2", "Make it tiny keeps the real task (open parent), no pile-up",
     "On a dreaded task choose 'Make it tiny' and do the 2-minute version. Then shrink the same task again to confirm pebbles do not accumulate.",
     "The dreaded task disappears and a 2-minute starter takes its place on Today, carrying an 'A tiny step toward X' eyebrow so the real task stays visible. Completing the starter does NOT mark the big task done: the spent starter is retired (no clutter) and the real task reappears with a warm nudge ('You started, that's the hard part. X is back when you're ready.'). Shrinking the same task repeatedly never piles up duplicate pebbles, only one is open at a time. Make it tiny again for the next step, or just complete the task.", "Both"),
    ("AI-09", "AI decompose", "P2", "Whole-task finish raises the held bloom (scaled)",
     "Finish the LAST step of a broken-down task. Try it on a long-lingering or chunky task, and separately on a small same-day one. Then repeat with Reduce Motion on.",
     "A warm radial bloom rises over a dimming scrim: 'You finished the whole thing', the task name in Newsreader italic, and a warm context line ('... since you first wrote it down. N small steps. All done.'). It holds longer and blooms larger for a long-dreaded or chunky task than for a quick same-day one. A tap dismisses it early, otherwise it auto-settles. Never confetti, points, or a number on screen. On Android the dimmed scrim is clean, with NO vertical pillar or banding behind it (the SVG background pools are disabled on Android, where they mis-render at large size). With Reduce Motion on, the held title and warm colour still show, only the movement is removed.", "Both"),

    # --- AI: Sort-for-me & Lighten today -------------------------------------
    ("AI-10", "AI decompose", "P2", "Edit or drop AI-suggested steps before accepting (review is un-losable)",
     "Break down a task (or run Chart a course). In the review: tap a step's TITLE and rewrite it; tap the small x on another; ALSO edit and remove a step under 'Then, as you get there' if phases are shown. Mid-edit, deliberately tap the backdrop and (Android) press back. Then accept.",
     "Tapping a title swaps it for an input (emptying reverts, never silently deletes). The x removes the row and 'Add N tasks' follows; zero steps disables Add. THE PLAN IS UN-LOSABLE: backdrop taps and the back button never dismiss the review (a stray tap mid-edit COMMITS the edit); only 'Add', 'Not these, start over', or the calm 'Not now' exit, and 'Not now' leaves the task untouched. Later-phase rows edit and remove exactly like main steps, and accepted phases mint the edited titles. Device check: a second title-edit opened mid-edit may revert (never lose) the first.", "Both"),
    ("AI-05", "AI triage", "P2", "Sort-for-me (triage + feedback)",
     "In the brain-dump type a MIXED pile, one per line (a couple of quick things, one that can wait, one big/vague). At one line a hint nudges 'one per line'; at two, 'Break it down' becomes 'Sort for me'. Run it.",
     "Shows a summary line ('Sorted: N for today, M for tomorrow, K to break down.'). Quick items stay on Today, can-waits move to tomorrow, big ones get an inline 'Looks big, break it down?' prompt. Calm, never scolding.", "Both"),
    ("AI-06", "AI lighten", "P2", "Lighten today (re-spread a full day)",
     "When today is heavy (6+ tasks, or 4+ on a low day), tap 'Lighten today'.",
     "The button only appears on a heavy day. It proposes re-spreading a few tasks to later days so today becomes doable, propose-then-accept. Free and ungated for everyone, never scolding.", "Both"),

    # --- Calendar -------------------------------------------------------------
    ("LB-01", "Calendar", "P1", "Open the calendar",
     "Open the Calendar. Browse to different days/months.",
     "A real Gregorian calendar. Navigation works, no crash.", "Both"),
    ("LB-02", "Calendar", "P1", "Completed tasks show on their day",
     "Complete a task today, open the Calendar on today.",
     "The completed task is listed under today.", "Both"),
    ("LB-03", "Calendar", "P2", "Old dreaded task is celebrated",
     "Complete a task that is old or high-complexity; view it in the Calendar.",
     "Marked 'a big one' / weighted celebration. Never shamed for being old.", "Both"),
    ("LB-04", "Calendar", "P2", "Scheduled tasks show on the calendar",
     "Defer a task to tomorrow (or use 'Date...'), then open the Calendar and tap that future day.",
     "The future day shows an outline marker; tapping it lists the task under 'Scheduled'.", "Both"),
    ("LB-05", "Calendar", "P2", "First-ever open is warm, not empty",
     "On a brand-new install with nothing ever completed, open the Calendar.",
     "The month reads 'This is where everything you finish will gather. Nothing yet, and that's a fine place to start.' (a welcome, never 'you did nothing'). No day shows a 'Nothing logged' line on this first run. Once anything is completed, the normal calendar behaviour returns.", "Both"),

    # --- Scrapbook ------------------------------------------------------------
    ("SB-01", "Scrapbook", "P1", "Make a scrapbook",
     "In a week with completions, tap 'Make a scrapbook'.",
     "Loading shimmer, then a still-life image + caption in the polaroid.", "Web"),
    ("SB-02", "Scrapbook", "P1", "Image surfaces the tasks",
     "Look at the generated image/caption for a known week.",
     "Objects evoke the actual tasks (e.g. laundry -> folded linen). No text in image.", "Both"),
    ("SB-03", "Scrapbook", "P1", "Finished list + 'a big one'",
     "Below the polaroid, read the 'This week you finished' list.",
     "All week's completed titles listed; big wins marked 'a big one'.", "Both"),
    ("SB-04", "Scrapbook", "P2", "Invite state",
     "Open a week that has completions but no scrapbook yet.",
     "Dashed frame + mauve '+', 'Turn this week into a keepsake', and the finished list still shows.", "Both"),
    ("SB-05", "Scrapbook", "P2", "Free-tier limit is graceful",
     "Generate a few scrapbooks in one day (free tier ~1-2/day).",
     "When exhausted, a calm wait/error. No crash, the holder stays intact.", "Web"),
    ("SB-06", "Scrapbook", "P2", "Scrapbook persists",
     "Make a scrapbook, restart the app.",
     "It is still there (device-local).", "Both"),
    ("SB-07", "Scrapbook", "P2", "A missing keepsake image degrades gracefully",
     "Open a week whose scrapbook image is gone (e.g. the R2 object was purged on an account delete while the local entry survived, or the stored image is corrupt).",
     "No blank polaroid. The week falls back to the calm 'That keepsake's picture isn't available anymore. Make a new one?' invite, and remaking overwrites it with a fresh image. Never a broken frame.", "Both"),
    ("SB-08", "Scrapbook", "P2", "Share the keepsake PAGE (caption in the pixels, never a link)",
     "Under a week's keepsake, tap 'Share this keepsake'. On Android pick an app from the system sheet; on the web try a browser WITH the Web Share API (Chrome/Edge/Safari) and one without (Firefox desktop). Also cancel the share sheet once. Inspect the shared/downloaded image itself.",
     "What is shared is the keepsake PAGE as one jpeg (doubledone-week.jpg): the image with its caption and a small 'DoubleDone · Week of {date}' line baked INTO the picture on a cream band, so no receiving app can strip the words (text attached beside an image is dropped by WhatsApp and friends, hence pixels). Never a public link, never loose message text, never raw task titles (the caption is the user-visible scene line). Same page on web (canvas) and Android (view-shot); if compositing fails the bare image still shares. BOTH stored image shapes must work on Android: an older local (data:) keepsake AND an R2-persisted https one (the 2026-07-12 'Sharing isn't available here' bug). On a Web Share browser the sheet opens; without it the jpeg downloads with one calm line, 'Saved. Share it anywhere you like.' Cancelling the sheet is quiet. A 'scrapbook.shared' event logs the how (shared / saved), never any content.", "Both"),

    # --- Auth & sync ----------------------------------------------------------
    ("AUTH-01", "Auth & sync", "P1", "Email sign-in (OTP)",
     "Sign in, enter your email, get the 6-digit code from your inbox, verify.",
     "Signed in. Settings shows 'Synced to <your email>'. (This emails your inbox.)", "Both"),
    ("AUTH-02", "Auth & sync", "P1", "Local tasks migrate on first sign-in",
     "Have some local tasks, then sign in for the first time.",
     "Local tasks sync into the account. None lost or duplicated.", "Both"),
    ("AUTH-03", "Auth & sync", "P1", "Sync across two devices",
     "Sign in on web and the Android build. Add a task on one.",
     "It appears on the other after sync.", "Both"),
    ("AUTH-04", "Auth & sync", "P2", "Last-write-wins, no dupes",
     "Complete / edit the same task on both devices close together.",
     "State converges. No duplicate rows, no flip-flop.", "Both"),
    ("AUTH-05", "Auth & sync", "P2", "Sign out",
     "Sign out from Settings.",
     "Returns to local/anonymous. No account data left visible.", "Both"),
    ("AUTH-06", "Auth & sync", "P2", "Offline then online",
     "Go offline, add/complete tasks, then reconnect.",
     "Changes sync up on reconnect. Nothing lost.", "Both"),

    # --- Account deletion -----------------------------------------------------
    ("DEL-00", "Account deletion", "P1", "PREREQ: create the delete function",
     "Run the delete_account() function from supabase/schema.sql once in the Supabase SQL editor.",
     "Function created. (One-time setup; cannot be rolled back.)", "Setup"),
    ("DEL-01", "Account deletion", "P1", "Delete account + data",
     "Settings -> Delete account and data -> confirm. (Use a throwaway account first.)",
     "Account and synced data are gone. Returns to a clean, signed-out Today.", "Both"),
    ("DEL-02", "Account deletion", "P1", "Originating device is wiped",
     "On the device you deleted from, look at Today, the Calendar, the scrapbook, and any routines.",
     "Nothing of the account remains locally: no tasks, an empty Calendar, no scrapbook, no routines. Only display prefs (theme, text size) persist.", "Both"),
    ("DEL-03", "Account deletion", "P3", "Known limit: second device",
     "On a second signed-in device after deletion, observe behaviour.",
     "It keeps local data until its next sync fails auth (documented limitation).", "Both"),

    # --- MCP server -----------------------------------------------------------
    ("MCP-00", "MCP", "P1", "PREREQ: copy your token",
     "Sign in, Settings -> AI agent access (MCP) -> Copy my token.",
     "Token copied (web) / shown selectable. Server URL visible.", "Both"),
    ("MCP-01", "MCP", "P1", "Connect a client",
     "Add the /mcp server to Claude Desktop via mcp-remote + your token (see docs/mcp.md).",
     "Client connects. Lists the full tool set: add_task, list_today, list_upcoming, complete_task, update_task, delete_task, break_down, search, fetch.", "Desktop"),
    ("MCP-02", "MCP", "P1", "add_task round-trip",
     "Ask the agent: add 'book the dentist' to my DoubleDone.",
     "Task appears in your Today (web/app) after sync.", "Desktop"),
    ("MCP-03", "MCP", "P1", "list_today",
     "Ask the agent: what's on my DoubleDone today?",
     "Returns your open tasks, each with an id. Repeating tasks due today appear once and drop after you tick or skip them.", "Desktop"),
    ("MCP-04", "MCP", "P1", "complete_task",
     "Ask the agent to complete one of the listed tasks by id.",
     "Marked done; reflects in the app after sync.", "Desktop"),
    ("MCP-10", "MCP", "P1", "add_task with a future due date",
     "Ask the agent: add 'call the plumber' for next Tuesday. PREREQ: MCP-01.",
     "The task is created dated (not on Today). It surfaces on that day, and appears in list_upcoming. Confirmation names the date. Sending both a due date and a repeat is refused calmly.", "Desktop"),
    ("MCP-11", "MCP", "P1", "add_task repeating",
     "Ask the agent: add 'water the plants' every 2 days (and separately, 'standup' every weekday). PREREQ: MCP-01.",
     "A repeating task is created (interval / weekly). It appears on the right days in the app and in list_today when due. A malformed repeat (e.g. weekly with no weekdays) returns a calm error, never a crash.", "Desktop"),
    ("MCP-12", "MCP", "P1", "list_upcoming look-ahead",
     "Ask the agent: what's coming up this week? Then: next 14 days. PREREQ: at least one future-dated and one repeating task.",
     "Lists future one-offs by date AND each repeat's next occurrence, in date order, each with an id. Window clamps to 1..30 days (default 7). A repeat already done for its next hit shows the following occurrence.", "Desktop"),
    ("MCP-13", "MCP", "P1", "update_task edits",
     "Ask the agent to rename a task, then to move its due date, then to make it repeat daily. PREREQ: MCP-01.",
     "Each change reflects in the app after sync. Setting a due date clears any repeat and vice versa. An update with nothing to change is refused calmly. Editing a missing id says so.", "Desktop"),
    ("MCP-14", "MCP", "P1", "delete_task is a recoverable tombstone",
     "Ask the agent to remove a task by id. PREREQ: MCP-01.",
     "The task disappears from the app (hidden), confirmed 'Removed.' It is soft-deleted (deleted_at set), never hard-deleted; the tombstone syncs across devices.", "Desktop"),
    ("MCP-15", "MCP", "P1", "break_down proposes, never adds (propose-then-accept)",
     "Ask the agent to break down a dreaded task (e.g. 'do my tax'). Read the steps. Then say 'yes, add them'. PREREQ: MCP-01.",
     "break_down returns small, ordered, time-boxed steps and adds NOTHING (it reminds you nothing was added yet). Only after you agree does the agent call add_task per step and they appear in the app. Steps read calm, no shame, no exclamation marks.", "Desktop"),
    ("MCP-16", "MCP", "P2", "break_down per-user hourly cap",
     "In a loop, ask the agent to break down tasks more than ~20 times within an hour. PREREQ: MCP-01.",
     "After the cap, break_down returns a calm 'Let's pause breaking things down for a bit' with NO AI spend, and recovers on its own the next hour. Other tools keep working throughout.", "Desktop"),
    ("MCP-17", "MCP", "P2", "ChatGPT Deep Research (search + fetch)",
     "Connect the DoubleDone connector to ChatGPT Deep Research (OAuth, MCP-07). Run a research prompt that references your tasks, e.g. 'summarise what's open in my DoubleDone about the move'.",
     "Deep Research calls search then fetch under the hood. search returns matching open tasks (by keyword, case-insensitive); fetch returns one task's detail. Results are read-only (nothing changes) and cite doubledone.app.", "OAuth"),
    ("API-01", "REST API", "P2", "Create a repeating task via POST /tasks",
     "In the Swagger UI (/api/v1/docs) or curl, POST /tasks with { title, repeat: { kind: 'weekly', weekdays: [1,3,5] } } and a bearer token.",
     "201 with the task carrying a normalised recurrence object + a 'repeats' summary ('Mon, Wed, Fri'), no due. Sending both due and repeat returns a calm 400, never a 500. An app-made and API-made repeat behave identically (shared cadence engine).", "API"),
    ("API-02", "REST API", "P2", "Search and look-ahead query modes on GET /tasks",
     "GET /tasks?q=dentist (substring search) and GET /tasks?upcoming=7 (next 7 days), with a bearer token.",
     "?q returns open tasks whose title matches (case-insensitive, open tasks only). ?upcoming returns future one-offs plus each repeat's next occurrence within the window (future only; today is /tasks?today). Precedence q > upcoming > today.", "API"),
    ("API-03", "REST API", "P2", "OpenAPI spec + Swagger UI current at 1.1.0",
     "Open https://api.doubledone.app/api/v1/docs and GET /api/v1/openapi.json.",
     "Swagger UI loads; the spec is version 1.1.0 and documents recurrence in the task shape, the Repeat input, and the ?q/?upcoming/?today query params.", "API"),
    ("MCP-07", "MCP", "P1", "Connect by URL (OAuth) from claude.ai / ChatGPT",
     "In claude.ai/Cowork: Settings -> Connectors -> Add custom connector, URL https://api.doubledone.app/mcp. In ChatGPT: Developer mode -> add connector, same URL, OAuth. Complete the DoubleDone sign-in: email -> 6-digit code -> Allow. Then ask the agent to add and list tasks.",
     "A DoubleDone sign-in page appears (warm paper, brand). Existing accounts only (an unknown email never receives a code). The consent screen names what the connector may do AND the host access is sent to. After Allow, add_task/list_today/complete_task work against your rows; the connection persists without re-pasting a token.", "OAuth"),
    ("MCP-08", "MCP", "P1", "Disconnect AI connectors (immediate kill switch)",
     "With a URL-connected agent active, in the app go Settings -> AI agent access (MCP) -> Disconnect AI connectors. Then have the agent try a tool.",
     "The app confirms 'Disconnected. Any connected agent must sign in again.' The very next tool call from the connector fails (must re-authorise); no ~1h lag. The pasted-token path is unaffected (it expires on its own).", "OAuth"),
    ("MCP-09", "MCP", "P2", "PKCE is required (OAuth 2.1)",
     "Drive /authorize with a registered client but WITHOUT a code_challenge (or with method=plain), e.g. via a crafted URL or the MCP Inspector.",
     "The request is rejected (invalid_request) before any sign-in UI or session is minted. Only S256 PKCE flows proceed. No grant is created.", "OAuth"),
    ("MCP-05", "MCP", "P2", "No-token gate",
     "In the MCP Inspector, call a tool WITHOUT the Authorization header.",
     "Calm 'Not connected' result (isError). Nothing in your account changes.", "Inspector"),
    ("MCP-06", "MCP", "P2", "Expired token re-copy",
     "Use a token older than ~1 hour, then re-copy a fresh one.",
     "Stale token fails cleanly; fresh token works. No data exposed.", "Desktop"),

    # --- Settings & comfort ---------------------------------------------------
    ("SET-01", "Settings", "P1", "Theme light / dark / system",
     "Switch theme between Light, Dark, and System.",
     "Applies immediately. Dusk palette correct in both; System follows the OS.", "Both"),
    ("SET-02", "Settings", "P2", "Text size small / default / large",
     "Change text size across all three.",
     "App scales. No clipping or broken layout at large.", "Both"),
    ("SET-03", "Settings", "P2", "Motion -> Reduce",
     "Set Motion to Reduce.",
     "Gentle fades and scrolling titles stop.", "Both"),
    ("SET-04", "Settings", "P2", "Privacy & data link",
     "Tap 'Privacy & data' in Settings.",
     "Opens the privacy screen.", "Both"),
    ("SET-05", "Settings", "P2", "Export your data",
     "In Settings -> Your data, tap 'Export your data' (works with no account).",
     "Web downloads a doubledone-export-<date>.json with your tasks + completions; native opens the share sheet. Tombstones excluded, completion data kept.", "Both"),
    ("SET-06", "Settings", "P2", "Send feedback in-app",
     "In Settings, tap 'Send feedback', type a note, tap Send.",
     "An inline box opens (no mail client, no leaving the app). Send shows 'Sending...' then a calm 'Thank you. It is on its way.', and the note arrives at the support inbox. On failure it shows a calm retry with the typed text kept. Needs the Worker deployed with /feedback + the FEEDBACK_TO secret set.", "Both"),
    ("SET-07", "Settings", "P2", "Daily reminder explains why it can't turn on",
     "In Settings, set 'Daily reminder' to On in a context where it can't be granted: deny the browser notification prompt on web, or have notifications blocked.",
     "The toggle returns to Off and a calm one-line reason appears under it (e.g. 'Notifications are off for DoubleDone. Turn them on in your settings, then try again.'), never a silent failure and never a raw error. Granting permission and retrying turns it On.", "Both"),
    ("SET-08", "Settings", "P2", "Daily reminder time picker",
     "Turn 'Daily reminder' On, then use the - / + stepper that appears under it to change the time. Step down toward 12:00 AM and up toward 11:00 PM (the ends disable). Reload and re-open Settings; also turn the reminder on from the Today footer and confirm it uses the chosen time.",
     "A '- TIME +' stepper shows ONLY while the reminder is On, with a calm 12-hour label (e.g. '9:00 AM'). Each tap moves the hour by one and the label updates instantly; the minus disables at 12:00 AM, the plus at 11:00 PM. The chosen time survives reload and the daily nudge fires at that hour (not a fixed 9am), consistently wherever the reminder is switched on. Free for everyone, no upsell.", "Both"),
    ("SET-09", "Settings", "P1", "Turn AI off, the app whole without it",
     "In Settings under 'AI', tap 'Turn AI off'. Then move through the app: capture, Today (long-press a task), the Menu, Calendar.",
     "Turning off is instant with a warm line ('AI is off. Everything stays on your device.'). Every AI affordance then disappears: Sort for me / Tidy this / Scan in capture; Break it down / Make it tiny / Combine / Plan my day on Today, including the per-task Break-down and Make-tiny; the 'Chart a course' menu entry (and the screen redirects away); and Calendar's scrapbook + weekly reflection. On-device Speak stays. Nothing calls the AI backend and the calm to-do loop still works fully.", "Both"),
    ("SET-10", "Settings", "P2", "Turn AI back on (informed consent)",
     "With AI off, in Settings tap 'Turn AI on', read the card, and tap 'Turn on AI' (also try 'Not now').",
     "Turning ON asks first: a card naming exactly what is sent and to whom ('sends the text you choose... to Anthropic's Claude... Nothing else ever leaves your device.'). 'Not now' cancels with no change. 'Turn on AI' restores every AI affordance across the app. The asymmetry is deliberate: off is one instant tap, on is a clear informed tap.", "Both"),
    ("SET-11", "Settings", "P3", "The installed version shows at the foot of Settings",
     "Scroll to the very bottom of Settings on Android and on web.",
     "One quiet line shows the installed version: 'v1.0.0 (20)' on Android (the number in brackets is the versionCode, the thing Play never shows a user), 'v1.0.0 (web)' on web. Exists because every store build shares the version NAME, so during the v19/v20 keyboard-fix confusion a tester had no way to say which build they held. When triaging any report, ask for this line first.", "Both"),
    ("I18N-01", "Languages", "P2", "App follows the device language (Italian / Spanish / French)",
     "Set the device (or browser) language to Italian, Spanish, or French and open DoubleDone fresh. Walk Today, the introduction, Settings, Calendar, capture, and close-the-day.",
     "The whole UI renders in that language: screens, buttons, hints, errors, notification copy, and screen-reader labels. Dates and times follow the region's convention (24-hour clock and day-month order where the locale uses them). Task titles the user typed stay exactly as typed (user data never translates), the brand stays 'DoubleDone', and legal pages stay English. Any missing string falls back to English, never a blank. An English device sees the app unchanged.", "Both"),

    # --- Accessibility --------------------------------------------------------
    ("A11Y-01", "Accessibility", "P2", "Screen reader (TalkBack)",
     "Enable TalkBack, navigate Today and capture.",
     "Controls are labelled. Dates read in a friendly way.", "Android"),
    ("A11Y-02", "Accessibility", "P2", "Touch targets",
     "Tap the small controls (done, chips, actions) with a thumb.",
     "Comfortable to hit. No fiddly mis-taps (hitSlop adequate).", "Both"),
    ("A11Y-03", "Accessibility", "P3", "Large text does not clip",
     "Set text size Large and walk every screen.",
     "Nothing important truncated or overlapping.", "Both"),

    # --- Theming / visual on device ------------------------------------------
    ("VIS-01", "Visual", "P1", "Native fonts render",
     "On the Android build, look at headers and body text.",
     "Serif (Newsreader) headers, Atkinson body, correct bold weights.", "Android"),
    ("VIS-02", "Visual", "P2", "Dark palette on device",
     "Switch to Dark on the Android build.",
     "Dusk dark, comfortable contrast, no harsh pure-black/white.", "Android"),
    ("VIS-03", "Visual", "P2", "Bold body text renders",
     "Find bold body text (e.g. emphasised labels) on Android.",
     "Bold reads as truly bold (the bodyBold fix), not faux/again-regular.", "Android"),

    # --- Reminders (Android) --------------------------------------------------
    ("REM-01", "Reminders", "P1", "Daily reminder fires",
     "Turn the daily reminder on (Settings > Daily reminder, or the Today footer), grant permission, and leave it on until the daily hour.",
     "A calm 'Your today is here when you are ready.' notification arrives at the scheduled hour (verified firing on a real Samsung device). Toggling it off cancels it.", "Android"),
    ("REM-02", "Reminders", "P2", "Reminder channel present",
     "Android Settings -> Apps -> DoubleDone -> Notifications.",
     "A DoubleDone reminder channel exists and is controllable.", "Android"),

    # --- Privacy --------------------------------------------------------------
    ("PRV-01", "Privacy", "P1", "Privacy policy reachable",
     "Visit doubledone.app/privacy (and via Settings).",
     "Loads, plain-English, matches the privacy posture.", "Both"),

    # --- Cross-platform / deploy ---------------------------------------------
    ("DEP-01", "Deploy", "P1", "Web loads + deep links",
     "Open doubledone.app, then hard-load /privacy and /sign-in directly.",
     "App loads; deep links resolve (SPA fallback), no 404.", "Web"),
    ("DEP-02", "Deploy", "P1", "Android APK installs + launches",
     "Sideload the latest APK and open it.",
     "Installs and runs. Core loop works.", "Android"),
    ("DEP-03", "Deploy", "P2", "Local-first offline",
     "Use core features (add/complete) with no network.",
     "Works offline; syncs later when signed in.", "Both"),

    # --- Premium (Stripe, test mode) -----------------------------------------
    ("PREM-00", "Premium", "P1", "PREREQ: Stripe test keys + webhook",
     "Set STRIPE_SECRET_KEY + STRIPE_WEBHOOK_SECRET as Worker secrets; register the /stripe-webhook URL in the Stripe sandbox.",
     "Worker deployed with the keys; webhook endpoint registered and reachable.", "Setup"),
    ("PREM-01", "Premium", "P1", "Paywall renders",
     "Open Settings -> DoubleDone Premium.",
     "The calm 'Keep every week' pitch, the A$5/mo price, the 1 -> 2 -> 4 tenure tiers. The feature list includes the personalisation pair (Quiet, the borderless look, and the seven colour themes) alongside Scan, Pin and the AI suite, so a user sent here from the Interface gate can see the thing they tapped.", "Both"),
    ("PREM-02", "Premium", "P1", "Free monthly gate routes to the paywall",
     "As a free user who already made a scrapbook this month, tap 'Make a scrapbook'.",
     "Routed to the Premium paywall, calm, never a shaming message.", "Both"),
    ("PREM-03", "Premium", "P1", "Checkout with a test card",
     "Signed in, tap Go Premium, pay with Stripe test card 4242 4242 4242 4242 (any future expiry, any CVC).",
     "Stripe Checkout opens, payment succeeds, returns to /premium?status=success.", "Web"),
    ("PREM-04", "Premium", "P1", "Entitlement flips to premium",
     "After the test checkout, reopen Premium and the Calendar.",
     "Premium shows active; the scrapbook is no longer monthly-gated (now weekly).", "Both"),
    ("PREM-05", "Premium", "P2", "Premium weekly wait stays calm",
     "As premium, make this week's allowance of scrapbooks, then try one more.",
     "A calm 'next ready in N days' message, never a paywall.", "Both"),
    ("PREM-06", "Premium", "P2", "Webhook rejects a bad signature",
     "POST a forged event to /stripe-webhook with no valid Stripe-Signature.",
     "400 bad signature; no entitlement change.", "Worker"),
    ("PREM-07", "Premium", "P1", "Webhook delivery succeeds (Stripe side)",
     "After the test checkout, open the Stripe event destination's delivery log.",
     "The checkout/subscription events show 'delivered' with a 200 from the Worker.", "Web"),
    ("PREM-08", "Premium", "P1", "Manage subscription opens the billing portal",
     "As premium: /premium -> Manage subscription.",
     "Redirects to the Stripe Billing Portal (cancel / update card / invoices).", "Web"),
    ("PREM-09", "Premium", "P1", "Cancel reverts to free",
     "In the portal, cancel immediately, then return to the app.",
     "Entitlement flips to free; the scrapbook is monthly-gated again; tenure (started_at) is preserved.", "Web"),
    ("PREM-10", "Premium", "P2", "Premium screen shows the renew / cancel date",
     "As premium, open /premium; then schedule a cancel-at-period-end in the portal and reopen.",
     "Reads 'Renews <date>' when active, and 'Premium until <date>, then free' when a cancel is scheduled.", "Web"),
    ("PREM-11", "Premium", "P3", "Comp allowlist: the owner email is always premium",
     "Sign in with the comp email (the owner account listed in server/src/comp.ts), without ever paying.",
     "Premium is active immediately: Settings and the Calendar show the scrapbook unlocked and Scan works, with no Stripe charge. A non-allowlisted free account stays free. The allowlist is checked against a cryptographically verified token on the costed gate, so a forged token cannot claim premium compute. REGRESSION CHECK (found on-device 2026-07-19): on /premium the panel reads 'You're Premium' and, because a comp has no Stripe customer, shows a calm 'nothing to manage here' line IN PLACE of a Manage button. It must never show 'Could not open the billing portal. Please try again.' (retrying a portal that does not exist). POST /portal answers such accounts 404 and the app reads that as nothing-to-manage, not a failure.", "Both"),
    ("PREM-12", "Premium", "P1", "Calendar insights: the premium 'Your patterns' card",
     "As premium (or with the dev Premium override on), finish a few tasks across a couple of days this week including one big/dreaded one, then open the Calendar and scroll below the Scrapbook. Tap 'Reflect on this week'.",
     "A calm 'Your patterns' card shows warm counts (finished this week, 'on N days', reclaimed old tasks named) with NO streak, score, percent, or 'missed' wording. 'Reflect on this week' returns one warm paragraph that only celebrates what was done (never a performance review) and changes nothing about your tasks. A 'lookback.summary.made' event is logged.", "Both"),
    ("PREM-13", "Premium", "P1", "Calendar insights: free sees a calm upsell, not a wall",
     "As a free user, open the Calendar and scroll below the Scrapbook, then tap the 'Your patterns' card.",
     "A calm one-line invite ('See what your weeks and months add up to'), never a teased count and never a wall. Tapping routes to /premium and logs 'premium.gate_hit' with reason 'insights'. The free user's calendar and their one monthly scrapbook are completely untouched.", "Both"),
    ("PREM-14", "Premium", "P2", "Premium: custom colour theme",
     "As premium (or with the dev Premium override on), open Settings -> Comfort -> 'Colour theme' and tap each of Dusk / Sage / Slate / Heather / Fog / Honey / Rose, reloading after choosing one, in BOTH light and dark mode. Then as a FREE user, tap a preset.",
     "Premium: tapping a preset repaints the WHOLE app in that palette (background, cards, accent, the brand) and the choice survives reload; the chosen preset is ringed; Dusk is the default and looks identical to the pre-themes app. Each preset has a tuned light and dark variant. Honey's buttons keep DARK labels (a calm gold can't carry white text). Free: the block shows a 'Premium' tag and tapping any preset routes to the paywall with no change applied. A lapsed subscriber keeps the preset they chose.", "Both"),
    ("PREM-15", "Premium", "P2", "Post-payment 'taking a while' recovery",
     "Complete a test checkout, then on the return /premium success screen simulate the entitlement being slow (e.g. the webhook delayed) so polling does not flip within ~10 tries.",
     "Instead of spinning forever it shows a calm message ('This is taking longer than usual. Your payment went through, give it a minute, then tap Refresh.') with a Refresh button and a pointer to send a note from Settings if it persists. Tapping Refresh re-checks and flips to premium once the entitlement lands. It never says the payment failed.", "Both"),
    ("PREM-16", "Premium", "P2", "Card-free 'Try Premium' one-month trial",
     "Signed in, on /premium tap 'Try Premium free for a month'. Confirm Premium unlocks. Tap it again on the SAME account. Advance the clock past 30 days (or inspect the trials row) and re-check entitlement. Also open the trial link signed OUT.",
     "First tap: a calm confirm, Premium turns on with NO card and NO Stripe (status 'trial'); the page shows 'Your free month', 'Free until <date>', and on web/Android the convert area shows the SAME monthly/annual toggle as the free paywall with the chosen price, and 'Go Premium to keep it' opens Stripe checkout for THAT plan (a trial member could only ever convert to monthly until 2026-08-01, when one asked to buy the year and could not; never 'Manage': a trial has no portal, so a Manage button here could only 404). On iOS the same trial copy shows with NO mid-trial CTA at all (StoreKit refuses a second purchase while premium and the trial never auto-charges; the 'Free until' line carries it), and never a Manage button. Second tap, same account: a gentle 'You've already had your free month' (never shame), no second trial granted. After expiry it reverts to free on its own, no charge ever. Signed out: an account is required (one trial per account), so the link routes to sign-in rather than granting. NOTE: this exact expected text was specified from the start, but the screen read the URL ?status= param instead of the entitlement status, so every trial user actually saw 'You're Premium' plus a dead Manage button until 2026-07-19. If this case regresses, that is the bug that came back.", "Both"),
    ("PREM-17", "Premium", "P2", "Annual vs monthly plan checkout",
     "On /premium use the Monthly / Annual toggle, then Subscribe. Confirm the Stripe Checkout reflects the chosen plan. Complete a test annual checkout. Then, as an already-subscribed user, hit Subscribe again.",
     "The toggle shows 'A$50/year, about two months free' for annual; Checkout opens the YEARLY price for Annual and the monthly price for Monthly, and the success path grants Premium either way. An already-subscribed user is refused a second Checkout (the server 409s and the app says 'You're already on Premium', never a double charge). Prerequisite: the Worker deployed with STRIPE_PRICE_ID_ANNUAL and that price live in Stripe.", "Both"),

    # --- Apple IAP (iOS) -----------------------------------------------------
    ("PREM-18", "Premium", "P1", "iOS paywall carries everything Apple requires",
     "On the iPhone (TestFlight), open Premium while signed out. Read the whole paywall.",
     "The paywall shows: the title 'DoubleDone Premium', BOTH prices from the store in the device's currency (A$5.00 / A$50.00 on an Australian storefront; StoreKit renders the viewer's own storefront price, so other regions see their converted price), the renewal line ('It renews on its own until you cancel it...'), a tappable Terms of use and Privacy policy, and a visible 'Restore a purchase'. There is NO mention of Stripe anywhere on the screen. The prices come from StoreKit, so if they read blank the offering is misconfigured (App Store Connect), not a code bug.", "iOS"),
    ("PREM-19", "Premium", "P1", "Anonymous Go Premium opens the purchase sheet (registration is OPTIONAL, App Review 5.1.1)",
     "Signed out on the iPhone (no account at all), open Premium, read the line under the button, then tap 'Go Premium' and complete a sandbox purchase. Kill and relaunch the app. THEN sign in with a fresh email + OTP.",
     "The button reads 'Go Premium' while anonymous and opens the StoreKit sheet directly, no sign-in wall anywhere (the wall was rejected under 5.1.1(v), 2026-07-28). The line under it explains, calmly: no account needed, Premium lives with the Apple ID, signing in any time brings it to other devices, and an existing web subscriber should sign in first. After the purchase Premium is ON while still anonymous, and SURVIVES a relaunch (the device's own RevenueCat entitlement is the source, no server row exists yet). Signing in later aliases the purchase onto the account (the webhook then writes D1), so it appears on the user's other devices: Apple's exact suggested flow.", "iOS"),
    ("PREM-20", "Premium", "P1", "Sandbox purchase unlocks Premium via the webhook",
     "Signed in with a sandbox Apple ID, tap Go Premium, complete the StoreKit sheet.",
     "The StoreKit sheet appears, the sandbox payment succeeds, the calm 'setting up' beat shows, then Premium flips on within ~20s (the RevenueCat webhook writes D1, the success-poll picks it up). Apple's accelerated sandbox clock then renews it several times and expires it, exercising the whole INITIAL_PURCHASE -> RENEWAL -> EXPIRATION chain against the real Worker. Prerequisite: the Worker deployed with RC_WEBHOOK_AUTH and the RevenueCat dashboard webhook set to the same secret.", "iOS"),
    ("PREM-21", "Premium", "P1", "A Stripe web subscriber is NEVER charged again on iOS",
     "As a user who already subscribed to Premium on the website (Stripe), install on iPhone, open Premium, and sign in with that same account.",
     "The screen shows the Premium (active) panel with NO buy button, so a second charge is impossible. Even if the buy button were reachable mid-load, the fresh entitlement read before buy() bails to the active panel. This is the double-charge guard, and the single most important iOS case.", "iOS"),
    ("PREM-22", "Premium", "P1", "Cancelling the StoreKit sheet shows nothing",
     "Tap Go Premium, then cancel/dismiss the Apple payment sheet.",
     "The app shows nothing at all: no error, no message, no toast. The user simply backed out. (Every OTHER failure does show a calm specific line: pending, already-owned, not-allowed, store-down.)", "iOS"),
    ("PREM-23", "Premium", "P1", "Restore works with no account at all, and always answers honestly",
     "Tap 'Restore a purchase' while signed out, on an Apple ID WITH a prior purchase. Then signed out on an Apple ID with NO purchase. Then repeat signed in.",
     "No sign-in is ever required (5.1.1: the receipt lives with the Apple ID, and the device's own entitlement now unlocks Premium locally). With a purchase -> 'Restored. Premium is back on.' and Premium turns on, account or not. With nothing -> 'Nothing to restore on this Apple ID...' It is NEVER a silent no-op (Apple rejects a restore that appears to do nothing).", "iOS"),
    ("PREM-24", "Premium", "P2", "Manage subscription opens Apple's sheet, not a browser",
     "As an Apple subscriber on the iPhone, open Premium and tap Manage subscription.",
     "Apple's own Manage Subscriptions sheet opens in-app. It never opens a browser or the Stripe portal.", "iOS"),
    ("PREM-25", "Premium", "P2", "An Apple subscriber on the web sees 'Apple handles it', not a portal error",
     "As a user whose Premium was bought on iPhone, open the web app, go to Premium, tap Manage subscription.",
     "The app says the subscription is managed in Apple's settings on the device it was bought on. It does NOT 404 the Stripe billing portal or show 'Could not open the billing portal' (the bug the entitlement source column exists to fix).", "Both"),
    ("SETL-01", "Settle", "P1", "The breathing room: rhythm, words, and the never-rules",
     "Open the day tools and tap Settle. Sit with it for a full minute. Toggle the guide pill off and wait ~90 seconds. Leave. Re-enter.",
     "The room holds ONLY: Leave (top-left), the breathing blob, one word slot, and the guide pill (reads 'Breathing guide · on/off'). NO title, NO timer or duration anywhere, NO stats, NO paywall or lock (free, unlimited, forever). The blob swells ~4s, holds ~1.5s, settles ~6.5s, on a loop, BRIGHTENING as it fills (an inner glow blooms at the top of the in-breath) and dimming as it empties; guide ON fades the three literal words in and out in phase ('in, gently…' / 'rest here' / 'slowly out…'), each breath wearing the next Dusk accent like Today's inscription, never a hard swap. Guide OFF: words stop, and at most one affirmation appears every 30-60s (seven lines rotating in order), fading in over ~3s and gone within 12, never stacking, never praising, its colour FROZEN for its whole display (never flipping when the breath advances the accent cycle). The toggle is REMEMBERED across visits. Leaving fades out on the out-breath (≤800ms) with 'Today is where you left it.', then Today exactly as left: no summary, no toast.", "Both"),
    ("SETL-02", "Settle", "P1", "Settle's two doors: the tools panel and the wind-down evening line",
     "Open the day-tools caret at any hour and read the order. Then after ~19:00, look under the wind-down line. Also check with AI off.",
     "The panel order is fixed forever: Plan my day, Focus on one thing, Lighten today, Settle, Close the day (AI off: Focus, Settle, Close). Settle renders as plainly as its neighbours, is NEVER the 'Right now' occupant at any hour, is never gated or locked, and the app NEVER opens it by itself. In the evening one soft italic line sits under the wind-down text ('If today is still loud, there's room to settle first.') and opens the room.", "Both"),
    ("SETL-03", "Settle", "P2", "The haptic breath and the reduce-motion room (device only)",
     "On a phone, enter Settle and hold the phone: feel for one light tap at the swell and two tiny taps ~300ms apart at the settle, nothing during the still. Turn the phone face-down and keep feeling. Then enable OS reduce-motion and re-enter.",
     "The haptic breath marks the phases and keeps working face-down and after the screen dims (no keep-awake, deliberately). Under reduce-motion the blob's GEOMETRY holds still and only its warmth breathes (or the room sits fully still), while the haptics and guide words carry the whole rhythm: the reduce-motion room is a designed variant, not a broken one. This is the ONE surface whose haptics do not silence under reduce-motion, by design.", "Both"),
    ("WN-01", "What's new", "P2", "The What's New card: once per announcement, never a modal, never to a fresh install",
     "On an existing (onboarded) install after an update that bumps the announcement id, open Today. Tap Got it. Reopen the app. Separately: complete onboarding on a FRESH install and check Today.",
     "A calm dismissible card in the coachmark's shape sits at the top of Today (NEVER a launch modal, never blocking capture): a small WHAT'S NEW title and one or two short lines. 'Got it' retires that announcement id forever (survives reload). A fresh install NEVER sees the card (onboarding stamps the current id: their whole app is new). A release with nothing user-visible bumps no id and shows nothing.", "Both"),
    ("WEB-02", "Web landing", "P2", "The landing page carries both store badges, web only",
     "As a fresh (not-onboarded) web visitor open doubledone.app and scroll to the closing section. Tap each badge. Also open the app on Android and iOS and confirm the landing never appears there.",
     "Official 'Download on the App Store' and 'Get it on Google Play' badges sit under the Begin button, visually equal in the row, with the Google-required trademark line in small print beneath. Each opens the correct store listing (country-less links, so every visitor lands on their own storefront in their own language). On native the landing redirects to Today before rendering, so the Play badge can never appear inside the iOS app (App Review would reject it). Screen readers hear the localised badge labels.", "Web"),
    ("ANA-01", "Analytics", "P2", "The Analytics Centre is token-gated, read-only, and answers the four questions", "",
     "GET /admin/analytics with no token -> 401 (or 503 if ANALYTICS_TOKEN is unset: an undeployed secret is never an open page). With ?token=<secret> -> a server-rendered Dusk page showing: premium counts by store and status plus active trials (Money), month-to-date AI spend with the month-end projection against the cap (AI spend), decompositions offered vs came-back-with-a-finished-step plus the median days to the first step (The moat), and keepsakes made in 28 days (Scrapbooks). No JavaScript, cache-control no-store, x-robots-tag noindex. Covered by analytics.test.ts; re-confirm once on the deployed Worker.", "Worker"),
    ("PREM-26", "Premium", "P1", "RevenueCat webhook rejects a wrong Authorization header", "",
     "POST /rc-webhook with a wrong Authorization header returns 401 and writes nothing to D1. With no RC_WEBHOOK_AUTH configured it returns 503. (Covered by revenuecat.test.ts; re-confirm on the deployed Worker.)", "Worker"),
    ("PREM-27", "Premium", "P1", "CANCELLATION keeps Premium on; EXPIRATION ends it",
     "Trigger (or replay) a RevenueCat CANCELLATION event for a subscriber, then an EXPIRATION.",
     "CANCELLATION leaves premium ON and the app reads 'Premium until {date}' (auto-renew off, access to period end), NOT an immediate revoke. A later EXPIRATION (with its date in the past) flips it to free and preserves the tenure start. A CUSTOMER_SUPPORT cancellation (a refund) DOES revoke immediately. (Locked by revenuecat.test.ts.)", "Worker"),
    ("PREM-28", "Premium", "P2", "A redelivered RevenueCat event is a no-op",
     "Deliver the same RevenueCat event id twice.",
     "The second delivery changes nothing (idempotent on the rc: namespaced event id, so it can never collide with a Stripe evt_ id). A dedup-store hiccup still writes (fail open), because the entitlement write is an idempotent upsert.", "Worker"),
    ("PREM-29", "Premium", "P1", "The Android AAB does NOT contain Google Play Billing",
     "Build the Android AAB, unzip it, and grep for com.android.billingclient. Also confirm the app still builds, installs, and Premium still sells via Stripe on Android.",
     "react-native-purchases (and Play Billing) is absent from the AAB: the expo.autolinking.android.exclude in client/package.json drops it. Android Premium is unchanged (Stripe Checkout in the browser). This is the regression guard on adding the native module.", "Android"),
    ("PREM-30", "Premium", "P1", "The web build is unchanged and still sells via Stripe",
     "Build the production web export (expo export --platform web). Open /premium on web.",
     "The export succeeds (the native purchases module never enters the web bundle, proving the purchases.ts / purchases.ios.ts split). The web paywall shows the catalog price, the 'Billed securely via Stripe' line, NO Restore control, and the Stripe checkout path is unchanged. The regression guard on the paywall rewrite.", "Web"),
    ("PREM-32", "Premium", "P1", "PRODUCTION paywall price matches App Store Connect",
     "After the app is RELEASED, install it from the App Store (not TestFlight) on an Australian account and open Premium. Read the monthly and annual prices.",
     "The paywall reads A$5.00 / month and A$50.00 / year, matching the App Store Connect price schedule for Australia, and matching what Apple's purchase sheet charges. CRITICAL: TestFlight and sandbox CANNOT verify this. They are known to answer product queries for the US storefront regardless of the account's real region, so a TestFlight paywall showing US dollars is expected and is NOT the bug. Only a production App Store install proves it. If production still disagrees with the sheet, that is a misleading pre-purchase price (an Australian Consumer Law problem, not just a display bug): pull the release via Pricing and Availability, Remove from sale. Prerequisite: app released. Gate the RELEASE on this, never the submission.", "iOS"),

    ("PREM-33", "Premium", "P1", "A dunning subscriber is never sold a second subscription",
     "Put a test Stripe subscription into past_due (in Stripe test mode: attach a failing card, or use the test clock to fail a renewal). Open /premium signed in as that user. Read the panel, then try to buy.",
     "Access has lapsed so the UPSELL panel shows, but with a calm notice at the top: 'Your Premium is paused while a payment retries. Updating your card brings it back. There is no need to buy again.' with an 'Update payment details' link that opens the Stripe portal (the customer id exists in this state). If the user taps Go Premium anyway, the server refuses with 409 billing_issue and the app explains the card is the fix, NEVER 'You're already on Premium' (false: they are not) and NEVER a second Checkout. REGRESSION GUARDS on the same server change (locked by stripe.test.ts): a LAPSED subscriber (status canceled, customer id kept) can still re-subscribe via a fresh Checkout, an abandoned checkout (incomplete) can still buy, and a trial still converts. The old guard keyed on premium && customerId, so a past_due user could mint a SECOND subscription whose webhook overwrote stripe_customer_id and orphaned the first: paying twice with one visible.", "Both"),

    ("PREM-31", "Premium", "P2", "Sandbox tester + Restore Behavior confirmed in the RevenueCat dashboard", "",
     "In the RevenueCat dashboard: Restore Behavior is set to keep-with-original (so TRANSFER never fires), the App Store Connect API key + In-App-Purchase key are uploaded, both products read 'Ready to Submit', and the IAPs are attached to the SAME App Store submission as the first binary (a first-time subscription submitted separately is not reviewed). A Sandbox Apple ID exists for testing.", "Setup"),

    # --- Quiet interface (Premium) -------------------------------------------
    ("QUI-01", "Interface", "P1", "Premium unlocks the Quiet interface; free is gated",
     "As premium (or with the dev Premium override on), open Settings -> Comfort -> 'Interface' and tap 'Quiet', then reload. Then as a FREE user, tap 'Quiet'.",
     "Premium: the whole Today surface goes borderless. Card outlines drop, and every action that looked like a button or a link renders as plain text with no outline or fill: the header rooms pills, 'Focus on one thing', 'Lighten today', 'Close the day', and 'Plan my day' (which in Standard is the filled premium gradient) all become plain text (accent for the AI / premium ones, soft ink for the calm structural ones). The energy / weight-of-today gauge is KEPT, not removed: it shrinks to a 3px hairline bar so the day's load still reads at a glance. The choice survives reload with 'Quiet' ringed. 'Standard' is the default and looks identical to the app today (gradient button, 6px gauge, outlined actions all return). Free: the block shows a 'Premium' tag and tapping 'Quiet' routes to the paywall with NO change applied ('appearance.locked' logged). A lapsed subscriber keeps whichever appearance they chose.", "Both"),
    ("QUI-02", "Interface", "P2", "Switching Standard <-> Quiet never moves the layout",
     "With a handful of tasks on Today (some open, one or two done), toggle Interface between Standard and Quiet a few times, watching the task rows and the capture line. On ANDROID, specifically do the full round trip (Standard -> Quiet -> Standard) from the Settings screen and then return to Today.",
     "Content stays put: no visible jump in row height or position as you toggle. Quiet removes chrome (row borders, fills, the per-task progress bar, the capture underline weight) but preserves spacing, so nothing shifts under the finger. On Android the round trip must leave rows pixel-perfect, in particular the row BOTTOMS never look clipped/'eaten' (regression: toggling while Today sat detached behind Settings in the native stack left rows at Quiet's slightly shorter measured height under Standard's styles; Today now remounts on an appearance change so native lays it out fresh).", "Both"),
    ("QUI-03", "Interface", "P2", "The held card is identical in Quiet and Standard (only the tint differs)",
     "Hold the SAME one-off task on Today in Quiet, then switch appearance and hold it again in Standard. Compare them side by side. Then in Quiet also hold a 'Later' task and (if you have one) a recurring task.",
     "BOTH appearances now give the same in-place card with the same actions in the same order (see TOD-07): there is no longer a Quiet model and a Standard model. Only the treatment differs: Quiet gets a soft wash with a rounded bleed, Standard the accent tint. Mark-as-a-lot and Pin act-and-dismiss (the row updates, the card closes). Pin shows only on Today one-off rows (not Later, not recurring), matching PIN-05; a lapsed subscriber still on Quiet who taps Pin is routed to the paywall, never wrongly pinned. 'Select more' closes the card and drops into multi-select with THIS task pre-selected ('1 selected'); Combine appears once a second task is selected. The one-time 'hold a task for more' coachmark in Quiet is a faint line with a small accent dot, not a filled pill.", "Both"),
    ("QUI-04", "Interface", "P2", "Quiet close-the-day reads as page text and stays reachable when full",
     "In Quiet, finish several tasks, then also test a very full day (roughly 16+ finished), and tap 'Close the day'.",
     "The wrap shows on the plain page colour with no dim scrim and no card outline, and 'Goodnight' is a plain accent text link, not a filled button. On a very full day the finished list still scrolls inside a bounded area and 'Goodnight' stays on screen and reachable, so the TOD-04d regression holds in Quiet too. Reopen / rested behaviour is unchanged.", "Both"),
    ("QUI-05", "Interface", "P3", "Quiet honours reduce-motion and every colour theme",
     "In Quiet, set Settings -> Motion -> Reduce (or the OS reduce-motion), then complete a task, open the held-state, and close the day. Separately, view Quiet across a couple of colour themes (e.g. Dusk and Sage) in BOTH light and dark.",
     "Nothing animates beyond the app's reduce-motion baseline: the held-state and the wrap appear without motion. Quiet derives its hairlines and washes from the ACTIVE palette, so it looks right on every colour theme in light and dark (not only Dusk), with no washed-out or invisible text.", "Both"),

    # --- Prioritise / pin a task (Premium) -----------------------------------
    ("PIN-01", "Pin", "P1", "Premium: pin a task as the day's one thing",
     "As premium (or with the dev Premium override on), tap-and-hold a one-off task to select it, then tap 'Pin' in the held row's actions.",
     "The task gets a calm mauve star and border and floats to the top of Today. 'Focus on one thing' then opens straight to it.", "Both"),
    ("PIN-02", "Pin", "P1", "Free: Pin routes to the upsell, never pins",
     "As a free user, select a single one-off task and tap the dimmed 'Pin' action.",
     "No task is pinned, and the Premium screen opens calmly (never a shaming wall). A 'premium.gate_hit' with reason 'pin' is logged.", "Both"),
    ("PIN-03", "Pin", "P2", "Only one task is pinned at a time",
     "As premium, pin task A, then hold task B and pin it.",
     "Task B is now pinned (starred, floated to the top), and task A is no longer pinned. At most one pin ever exists.", "Both"),
    ("PIN-04", "Pin", "P2", "A pin syncs across devices",
     "As premium and signed in, pin a task, then sign out and back in (or open a second device).",
     "The task is still pinned after the sync round-trip (the star and float persist). Needs the pinned_at column applied in Supabase.", "Both"),
    ("PIN-05", "Pin", "P3", "Pin is offered only on one-off Today rows",
     "Hold a recurring task, then hold a task in the 'Later' list.",
     "No 'Pin' action appears on a recurring task's held card, and a Later task's card carries no Pin either (pinning is Today-only and one-offs only).", "Both"),

    # --- Energy matching (freemium) -------------------------------------------
    ("NRG-01", "Energy", "P1", "What fits right now: one question, one pick, propose-only",
     "With AI on and 2+ open one-off tasks, tap 'Focus on one thing'; in the 'Which one?' picker, 'What fits right now?' sits above the task choices in a soft hairline pill (its border makes it read as tappable; it lives INSIDE Focus, not as a standalone Today button). Tap it, pick an energy level (Running low / Somewhere in between / Feeling good). Then tap 'Start with this'; separately try 'Not now'.",
     "A calm one-question card with the privacy line ('Your open tasks go to Claude to pick one. Nothing is added or changed.'). Choosing a level shows ONE task from today's list with a short warm line on why it fits (never shaming, never about what's undone). 'Start with this' opens Focus mode on exactly that task; 'Not now' just closes, nothing changed. A failed call shows a calm 'Couldn't find a pick just now.' line and does NOT spend a free pick. The entry hides when AI is off or fewer than 2 tasks are open. Needs the Worker deployed with /energy.", "Both"),
    ("NRG-02", "Energy", "P1", "The freemium meter: 15 a month, reminders at 10 and 5, then the paywall",
     "As a FREE user, use energy matching repeatedly in one calendar month (seed doubledone.energyUses.v1 to fast-forward). Watch the 10-left and 5-left moments, then the 16th tap. Then as premium, use it well past 15.",
     "Each successful pick spends one of 15 free picks a calendar month (counted locally, like the scrapbook gate). Exactly at 10 left and at 5 left, the result shows one calm line: '10 of your 15 free picks left this month.' (then 5). Past 15, tapping the entry routes to the paywall (which lists 'Energy matching without limits'), never a shaming wall, and no AI call is made ('premium.gate_hit' reason 'energy' logs). The counter resets on the 1st of the month with no carry-over. Premium is unlimited and never sees a counter line.", "Both"),

    # --- Scan a list (OCR, Premium) ------------------------------------------
    ("OCR-01", "Scan", "P1", "Premium: the Scan button opens the camera",
     "As premium, open the add panel and tap the Scan (camera) pill beside Speak.",
     "The camera screen opens: a live viewfinder with a shutter and a Photos shortcut on a device, or a 'Choose a photo' prompt on web. No paywall.", "Both"),
    ("OCR-02", "Scan", "P1", "Free: Scan routes to the upsell, never opens the camera",
     "As a free user, open the add panel and tap the Scan pill.",
     "Routed to the Premium screen calmly (never a wall). A 'premium.gate_hit' with reason 'ocr' is logged. No camera opens.", "Both"),
    ("OCR-03", "Scan", "P1", "Photograph a list and the tasks land in the box",
     "As premium on a device, tap Scan, photograph a short printed or handwritten list (fill the frame), wait for 'Reading your list...'.",
     "The tasks it reads appear in the brain-dump box, one per line, editable. Nothing is auto-added to Today, and tapping Add commits them. An 'ocr.captured' event is logged.", "Device"),
    ("OCR-04", "Scan", "P2", "Gallery fallback reads an existing photo",
     "Tap Scan, then 'Photos' (device) or 'Choose a photo' (web), and pick a photo of a list, a note, or a whiteboard.",
     "Same result: the read tasks land in the box for review. Covers a screenshot of a texted list, or a photo taken earlier.", "Both"),
    ("OCR-05", "Scan", "P2", "An unreadable photo fails calmly",
     "Scan a blank, blurry, or list-free image.",
     "A calm line ('I couldn't read any tasks from that. Try again...'), the camera stays open, never a crash and never a shaming message.", "Both"),
    ("OCR-06", "Scan", "P2", "Camera denial is never a dead end",
     "On a device, deny the camera permission when prompted.",
     "A calm screen offers 'Allow camera' and 'Choose from photos instead', and the gallery path still reads a list.", "Device"),
    ("OCR-07", "Scan", "P3", "AI egress is disclosed at the point of use",
     "Open the Scan screen (device or web).",
     "The note 'Your photo is sent to the AI to read your list, then discarded. It is never stored.' is visible. The D1 'ocr' telemetry row holds only the image size and task count, never the image or the titles.", "Both"),

    # --- Chart a course (Premium) --------------------------------------------
    ("CHART-01", "Chart", "P1", "Premium: chart a course toward a goal",
     "As premium (or dev Premium override on), open Rooms, tap 'Chart a course', type a goal like 'get fit for a 10k', tap 'Suggest steps'.",
     "A calm one-line heading plus 3-7 ticked next steps appear, and nothing is added yet. Tapping 'Add N tasks' lands them on Today (the first undated, later ones spread forward), each an ordinary task, and returns to Today. The 'chart.requested' then 'chart.added' events are logged.", "Both"),
    ("CHART-02", "Chart", "P1", "Free: charting routes to the upsell, never plans",
     "As a free user, open Rooms, tap 'Chart a course', type a goal, tap 'Suggest steps'.",
     "Routed to the Premium screen calmly (never a wall), and a 'premium.gate_hit' with reason 'chart' is logged. No plan is generated and nothing is added.", "Both"),
    ("CHART-03", "Chart", "P2", "Propose-then-accept: nothing auto-adds",
     "As premium, generate a plan, untick two steps, then tap 'Add'. Separately, generate a plan and back out with 'Not these, start over' or Back.",
     "Only the ticked steps are added as plain tasks, and backing out adds nothing. Today was unchanged before accepting.", "Both"),
    ("CHART-04", "Chart", "P2", "A goal that cannot be mapped fails calmly",
     "Enter an empty goal (the button is disabled), then a nonsensical goal and submit.",
     "The empty case cannot submit. A nonsensical goal shows one calm line ('I couldn't map that out just now'), the goal stays editable, never a crash or a shaming message.", "Both"),
    ("CHART-05", "Chart", "P2", "A deadline spreads the steps across the timeframe",
     "As premium, type a goal, tap a 'By when?' chip (e.g. 'In 2 months'), then Suggest steps and Add.",
     "The steps are paced for that timeframe, and the accepted tasks spread from Today out to the chosen date (not crammed into the next few days). 'No deadline' keeps the gentle one-per-day default.", "Both"),

    # --- Plan my day / sequencing (Premium) --------------------------------
    ("SEQ-01", "Sequence", "P1", "Premium: Plan my day suggests a calm sequence",
     "As premium with 3+ open one-off tasks on Today, tap 'Plan my day'. Answer the questions it asks first (or skip them).",
     "A proposal card lists today's tasks in a suggested order, each with a short calm reason. Nothing reorders until 'Use this order' is tapped, then the list re-sequences in place (no dates change, no task moves to another day). A 'sequence.accepted' event is logged.", "Both"),
    ("TOD-27", "Today", "P1", "The constant frame: one fixed action layer, a clock-set slot, Energy pills",
     "Open Today in the MORNING (before 11), around MIDDAY (11-17), and in the EVENING (17+), reading the 'RIGHT NOW' slot above the Add bar each time. Tap the caret next to it. Add and complete tasks and watch whether ANY control moves. Tap an unavailable (faded) tool. Set the Energy pills to High on a 7-task day and re-open the caret; set Low and read the gauge. Reload the app and check the pill. Turn AI off in Settings and re-open the caret. Also check with 0 tasks, and in select mode, and after closing the day.",
     "ONE fixed layer sits above the Add bar, always the same pixels: an overline 'RIGHT NOW' + the tool that suits the HOUR (morning: Plan my day; midday: Focus on one thing; evening: Close the day), plus a 44px caret ('All day tools'). The occupant is set by the CLOCK at open and never changes mid-session; it NEVER shows Lighten today. The caret opens a panel growing upward with the SAME four tools in the SAME day order every time: Plan my day, Focus on one thing, Lighten today, Close the day, the current occupant tagged 'now'; behind it the list dims 40% and never moves; tapping the dim closes. NOTHING in this layer appears, vanishes or slides as tasks are added or completed: an unavailable tool stays in place at lowered contrast (no lock, no strikethrough) and tapping it explains in one plain line what the TOOL needs ('Plan my day needs two or more tasks.' / 'Lighten today is for a full day.'), never what the person lacks. The Energy pills (Low/Normal/High) sit under a small 'ENERGY' overline with the gauge at the top, and the WHOLE panel (gauge + label + pills) renders on EVERY open day, including an empty one, where the gauge reads 'A clear day.' (real-user report 2026-07-26: without the gauge the pills floated headerless and read as lost text). A fresh day defaults to Normal. High re-scales what 'full' means (7 tasks stop being heavy, so Lighten goes quiet until ~9), Low IS the low-capacity day (the old toggle is gone; the same warm affirm appears), and the choice survives a reload but RESETS to Normal next morning (a day-state, never a setting). With AI OFF the panel holds only Focus and Close (AI tools are gone, not faded, per the AI-off rule). The old pile is gone: no Focus button above the list, no 'Today's looking full' line, no stacked Lighten/Plan buttons, no separate wind-down line (in the evening the italic inscription under Add becomes the wind-down sentence), and Close the day is never buried under a long list again. The layer hides in select mode and on a closed day.", "Both"),
    ("TOD-28", "Today", "P2", "The rested screen mentions the scrapbook once, at the earned moment (the ladder's last rung)",
     "With AI on, the reminder and widget asks already spent (or reminder on, widget placed), NO scrapbook ever made, and at least 3 finishes this week, close the day and read the rested screen. Tap 'See your Calendar'. Then close a day again on another evening. Separately: repeat with AI off, and repeat with only 1-2 finishes.",
     "One calm mention appears under the goodnight text: 'You've finished enough this week to make a keepsake of it. It's waiting in your Calendar.' with 'See your Calendar' and 'Not now'. Either answer spends the mention FOREVER (it is the third and LAST rung of the one-ask-at-a-time ladder; reminder and widget always win the evening if pending). 'See your Calendar' opens the Calendar where the scrapbook lives. It never shows with AI off, never with fewer than 3 finishes this week, never to someone who has already made a scrapbook, and never twice.", "Both"),
    ("SEQ-07", "Sequence", "P1", "Plan my day asks about the day BEFORE it sorts",
     "As premium, tap 'Plan my day'. Read the sheet that opens. Tap an answer, then tap the SAME answer again. Try sorting with all three answered, with only one answered, and with none answered ('Sort today' straight away). Also tap 'Not now'. Separately, as a FREE user, tap 'Plan my day'.",
     "Before any sorting, a calm sheet asks three OPTIONAL questions: 'How's your energy?' (Running low / Somewhere in between / Feeling good), 'What kind of day?' (Work day / Day off) and 'Where are you?' (Indoors / Out and about / Either), under 'Answer what you like. Skip what you don't.'. A chosen answer tints (accent border, fill and bold accent text); tapping it AGAIN clears it, so a mis-tap is undone the way it was made. Skipping a question sends nothing for it, so it never becomes an assumption, and skipping all three sorts exactly as it always did. The order that comes back reflects what was said (a low-energy day starts small; a day off does not front-load work; 'out and about' clusters the errands), but the app NEVER comments on how you feel. 'Not now' closes with nothing sent and the day untouched. A FREE user hits the Premium screen BEFORE the questions, never after answering them. NEW (the constant frame): when the Energy PILLS at the top of Today were touched today, the sheet's energy question DOES NOT APPEAR at all, and the pill's answer rides the request instead (Low->low, Normal->medium, High->good): energy is read, never asked twice. Only when the pills are untouched does the sheet ask. NOTE: there is deliberately no weather question. A model has no live weather data, and a forecast would need an API plus a location permission; 'Where are you?' is what the sort actually needs.", "Both"),
    ("SEQ-08", "Sequence", "P1", "The suggested order can be edited before it is used",
     "Get a proposal from 'Plan my day'. Use the up and down arrows on a row. Try the up arrow on the FIRST row and the down arrow on the LAST. Remove a row with x. Remove every row. Then tap 'Use this order'.",
     "Each row carries up / down / x. Up and down swap the row with its neighbour immediately, renumbering as they go; the first row's up and the last row's down are visibly dimmed and do nothing (they are disabled, not hidden, so rows never change shape as things move). 'x' takes that task OUT OF THIS PLAN ONLY: the task stays on Today untouched, is not deleted, and is not moved to another day. Removing everything is allowed and says 'Nothing left in the plan. That's allowed.' with the accept button gone (an empty plan cannot be applied). Nothing at all changes on Today until 'Use this order' is tapped; 'Not now' or closing discards every edit. Editing is by buttons, never drag: dragging is unusable with a screen reader and hard with shaky hands.", "Both"),
    ("SEQ-02", "Sequence", "P1", "Free: Plan my day routes to the upsell, never reorders",
     "As a free user with 2+ tasks on Today, tap 'Plan my day'.",
     "The Premium screen opens calmly (never a wall), a 'premium.gate_hit' with reason 'sequence' is logged, and the day's order is unchanged.", "Both"),
    ("SEQ-03", "Sequence", "P2", "'Not now' leaves the day untouched",
     "As premium, open the proposal, then tap 'Not now' or the backdrop.",
     "The order is exactly as before, nothing reordered, and no manualOrder is written.", "Both"),
    ("SEQ-04", "Sequence", "P2", "An accepted order survives a reload (local-first)",
     "As premium, accept an order, then fully reload the app.",
     "Today still shows the accepted order after reload (manualOrder persists on-device). Note: the order does not yet sync across devices, which is a documented follow-up.", "Both"),
    ("SEQ-05", "Sequence", "P3", "A pinned task still wins the very top",
     "As premium, pin a task, then accept a 'Plan my day' sequence that puts a different task first.",
     "The pinned task stays at the very top, and the accepted order applies to everything below it.", "Both"),
    ("SEQ-06", "Sequence", "P2", "Plan my day offers to lighten a heavy day",
     "As premium on a heavy day (6+ tasks), tap 'Plan my day', accept the order.",
     "After the order applies, a 'Still a full day?' card offers to push a few tasks out to later days. Yes runs the re-spread (propose-then-accept), No leaves it ordered. On a calm day the offer never appears.", "Both"),

    # --- Big task (free, all tiers) -------------------------------------------
    ("BIG-01", "Big task", "P1", "Mark a task big: the tag, and the weight it adds",
     "On Today with a few one-off tasks, tap and hold one, then tap 'Mark as a lot' in the held row's actions. ALSO hold a RECURRING task and confirm 'Mark as a lot' is offered there too (a chore can be a lot; this used to be offered only via the bulk bar).",
     "The card closes and the task shows a calm accent 'Big' tag beside its title (never red, never a warning), with a brief validating note ('Marked as a lot...'). The weight gauge fills further and reads heavier (one big task floors the bar to at least 'A full day, but doable.'). Free for everyone, no upsell.", "Both"),
    ("BIG-02", "Big task", "P2", "Big is a multi-select toggle",
     "Hold a task, 'Select more', tick another, then tap 'Mark as a lot'; then re-select tasks that are all big and tap the action again (it now reads 'Not big').",
     "The first tap marks every selected task big at once. When all selected tasks are already big the action reads 'Not big' and clears the mark off all of them. A mixed selection marks all big (the additive default).", "Both"),
    ("BIG-03", "Big task", "P2", "A lone big task lifts the bar but is not a re-spread problem",
     "Have just one or two tasks on Today, one marked big.",
     "The weight gauge reads at least 'A full day, but doable.' (the big task is felt), but 'Today's looking full' and 'Lighten today' do NOT appear for a lone big task (re-spreading cannot dissolve one big rock; Break it down is the tool). A big task plus a real pile (weighted load 6+) does surface Lighten today.", "Both"),
    ("BIG-04", "Big task", "P2", "Finishing a big task is a big-win in the Calendar",
     "Mark a task big, complete it, then open the Calendar to its day.",
     "The completed task carries the warmer 'a big one' treatment on its day, the same payoff a long-dreaded or chunky task earns. The big flag never adds disappointment if the task is left unfinished.", "Both"),
    ("BIG-05", "Big task", "P2", "Big syncs across devices (and old marks survive the upgrade)",
     "Signed in on two devices (e.g. web + Android). Mark a task big on device 1, wait for or trigger a sync, then open device 2. Then unmark it on device 2 and check device 1. Prerequisite: an account with the same task on both.",
     "The Big tag appears on device 2 (with its weight-gauge effect), and the unmark travels back: no resurrection, the newest change wins. Marks made BEFORE this build (when big was device-only) are seeded up to the account on that device's first sync rather than erased. One known transition quirk: a device still on a PRE-sync build can re-surface a cleared big once when it upgrades; clearing it again sticks everywhere. Manual order (Plan my order) remains per-device by design.", "Both"),

    # --- Navigation / responsive ---------------------------------------------
    ("NAV-01", "Rooms", "P3", "Rooms sheet caps its width on wide web",
     "On a wide desktop browser, open Rooms (the header pill).",
     "The sheet is a centred column (about 560px, matching the page content), not full-bleed, so the 'Premium' gradient pill on 'Chart a course' sits beside its label rather than at the far screen edge. On a phone the sheet stays full-width with the pill at the row's edge.", "Web"),

    # --- Ours: a shared list between two people --------------------------------
    # Needs TWO accounts, both on the ours_allowlist, and supabase/ours.sql applied.
    ("OUR-01", "Ours", "P1", "The door only exists for an allowed account",
     "Sign in as an account that is NOT on the ours_allowlist and open Settings. Then sign in as one that is.",
     "The 'Ours' row is absent for the first account and present for the second. Nobody is ever offered a door that answers 'shared lists aren't open yet'.", "Both"),
    ("OUR-02", "Ours", "P1", "Signed out: the explanation, never a nag",
     "Sign out, then open /ours directly.",
     "A calm one-screen explanation of why sharing needs an account, and a way to sign in. No pressure, no count of what you are missing.", "Both"),
    ("OUR-03", "Ours", "P1", "Name the list and mint a code",
     "Ours > Start a shared list. Pick 'The house'. Type your own name. Type the OTHER account's email exactly. Get a code.",
     "A six-character code appears grouped as XXX-XXX, containing no 0, 1, I, L or O. The screen says it works once, for the next day, and only for the address you gave.", "Both"),
    ("OUR-04", "Ours", "P1", "Join from the second account",
     "On the second account, Ours > Join with a code. Type the code (try it lower-case and with the dash). Type your own name. Join.",
     "You land on the shared list showing the FIRST person's chosen name and the list's name ('The house'). No email address is shown anywhere, to either person.", "Both"),
    ("OUR-05", "Ours", "P1", "The first person sees the arrival without refreshing",
     "Leave the first account sitting on the Ours screen while the second joins.",
     "Within about ten seconds it announces that they joined, with a way out for someone who meant a different person. It never says how long they took.", "Both"),
    ("OUR-06", "Ours", "P2", "A wrong code says one calm thing",
     "Try a made-up code, then a code meant for a different address, then the same code twice.",
     "Every one gives the SAME single line ('That code isn't valid.'). Nothing hints at which reason, so a guesser learns nothing. Repeated wrong tries eventually ask you to have a break.", "Both"),
    ("OUR-07", "Ours", "P2", "Mistyped invitee address is recoverable",
     "Mint a code with a deliberately wrong email, then use 'Wrong address? Start again' and mint a fresh one with the right address.",
     "The new code works and the old one does not. Nothing had to be deleted to recover, and no address is ever displayed back to you.", "Both"),
    ("OUR-08", "Ours", "P2", "Rename the list, and both sides see it",
     "On the shared list, tap the list's name, type a new one, and confirm. Reopen Ours on the other account.",
     "Both people see the new name. An emptied name falls back to the app's own word, in each person's own language.", "Both"),
    ("OUR-09", "Ours", "P1", "Leaving freezes and takes nothing",
     "On one account, Leave this list. Then open Ours on BOTH accounts.",
     "Both see 'This list is closed' and can still read everything. Nothing was deleted from either side. Neither person is told off, and neither is told what the other did.", "Both"),
    ("OUR-10", "Ours", "P2", "Remove a closed list from your side only",
     "On a closed list, tap 'Remove this list'. Then check the other account.",
     "It disappears for you. The other person still has their copy and is not notified.", "Both"),
    ("OUR-11", "Ours", "P1", "Nothing anywhere attributes a task to a person",
     "With a shared list live, look at Ours, Today, the Lookback and Settings on both accounts.",
     "No initials, no 'done by', no counts of who did what, no progress bar across two people, no activity feed. The data to build one does not exist.", "Both"),
    # The six that the Phase 3 audit found blocking a dogfood. Every one of these is a path a
    # tester walks in the first ten minutes, which is why they are all P1.
    ("OUR-13", "Ours", "P1", "The code stays on screen from the moment it is minted",
     "On a throttled connection (devtools, slow 3G), fill the create form and tap the code button.",
     "The six characters appear as soon as the button stops spinning and are NEVER replaced by the intro screen. The server returns that code exactly once, so a flash back to 'Start a shared list' loses it for good.", "Both"),
    ("OUR-14", "Ours", "P1", "'Get a new code' reaches the form and mints a different code",
     "From the waiting screen, tap the get-a-new-code action. Change the email. Submit.",
     "The create form appears with your name and list name already filled. Submitting shows a DIFFERENT six characters, and the old one no longer works. Your chosen name is unchanged afterwards.", "Both"),
    ("OUR-15", "Ours", "P1", "Leave a list, then start a new one WITHOUT deleting the old",
     "Leave a shared list. On the closed screen, start a new shared list with someone else.",
     "Starting a new list is offered on the closed screen itself and works. The old list is NOT deleted and is still readable. At no point is destroying it the only way forward.", "Both"),
    ("OUR-16", "Ours", "P1", "A failed read never offers to start a list you already have",
     "With a live shared list, turn the network off and open Ours.",
     "A calm connection line. The screen never says you have no shared list and never offers to start one. Turning the network back on recovers without a restart.", "Both"),
    ("OUR-17", "Ours", "P2", "A list nobody joined can be left",
     "Create a list, get a code, give it to nobody. Return to Ours.",
     "There is a way out on the waiting screen, and its wording does not claim it closes for two people. After leaving, you can start or join a list again.", "Both"),
    ("OUR-18", "Ours", "P2", "The arrival announces once and stays announced",
     "Sit on the waiting screen while the other account joins. Background the app and return.",
     "The name appears once and does not flicker back to waiting and re-announce. The way out beside it ('that wasn't who I meant') never appears on a list you did not just gain a person on.", "Both"),
    ("OUR-19", "Ours", "P2", "A repeat the app cannot read is SHOWN, never hidden",
     "Requires two app versions. On the newer one, set a shared task to a cadence the older build does not have. Open the list on the older build.",
     "The task is visible on both, with whatever cadence line the newer app left. It is never silently absent from one person's list, and the older app does not erase the schedule when it syncs.", "Both"),
    # --- The room, the bridges, repeating, the guards (round two) ---------------
    ("OUR-20", "Ours", "P1", "The room opens from Today and from the Menu",
     "With a live shared list, tap the 'Ours \u00b7 {name}' row on Today. Then go back and reach it through Menu \u203a Ours.",
     "Both land on the shared list itself, not the pairing screen. The header line reads 'Kept with {their name}' and tapping it goes to the pairing/management screen.", "Both"),
    ("OUR-21", "Ours", "P1", "Add, tick and UN-tick on the shared list",
     "Add a task in the room. Tick it. Un-tick it. Watch the other account (reload or wait ~15s).",
     "Every step appears on both. Un-ticking works as easily as ticking, on both sides. No affirmation or celebration fires on Ours.", "Both"),
    ("OUR-22", "Ours", "P1", "Un-tick a one-off the DAY AFTER it was ticked",
     "Tick a shared one-off. Change the device date to tomorrow (or wait a day), reopen the room, tap it again.",
     "It un-ticks. (The pre-fix build could never un-tick it again, on any device, forever.)", "Both"),
    ("OUR-23", "Ours", "P1", "Bring to my Today makes a COPY, and marks only yours",
     "Hold a shared row, tap 'Bring to my Today'. Look at your Today, then at the shared list on BOTH accounts.",
     "Your Today shows the task with a faint '\u00b7 Ours'. The shared row is completely unchanged on both sides, with no marker of any kind. The held card now reads 'Already on your Today' and cannot make a second copy.", "Both"),
    ("OUR-24", "Ours", "P1", "Your tick on the copy closes both",
     "Tick the brought copy on your Today. Check the shared list on the other account.",
     "The shared row is done. It enters YOUR Lookback. Nothing on the other side says who did it.", "Both"),
    ("OUR-25", "Ours", "P1", "Ticking a brought copy on a device that has NEVER opened the room",
     "Bring a task over on phone A. Sign in on a second device (or clear its local data) and go straight to Today WITHOUT opening Ours. Tick the copy there.",
     "The shared row closes. (The pre-fix build silently discarded the tick with no error.)", "Both"),
    ("OUR-26", "Ours", "P1", "Their tick leaves a rest-note, never a strike-through",
     "Bring a task over. On the OTHER account, tick the shared row. Return to your Today.",
     "Your copy is gone from the day, replaced by a dashed sage line: 'Handled on Ours. It's off your day.' It is NOT struck through and does NOT appear in your Lookback. Leaving Today and coming back clears the note.", "Both"),
    ("OUR-27", "Ours", "P2", "Share to Ours warns about the length BEFORE it copies",
     "Hold one of your own Today tasks with a very long title (500+ chars), open More, tap 'Share to Ours'.",
     "A line says Ours keeps titles shorter and this one will be trimmed, BEFORE the copy is made. The copy appears on the shared list. 'Share to Ours' is absent entirely when there is no live list.", "Both"),
    ("OUR-28", "Ours", "P1", "A repeat set on Ours uses the SAME cadence sheet as home",
     "Hold a shared row, tap 'Repeat\u2026'. Compare with Menu \u203a Repeating \u203a Edit on a personal task.",
     "Identical controls. The commit button NAMES the cadence ('Every day', 'Every 2 days'), on both. The Ours sheet adds one line: 'You'll both see it on its day.' It appears on its days for both people.", "Both"),
    ("OUR-29", "Ours", "P1", "A cadence this build cannot read is inert, and says why",
     "Get a shared row carrying a recurrence this build does not understand (a newer build, or hand-write one via the API).",
     "The row is SHOWN, never hidden. Tapping it does nothing and a line explains it is safe and will appear on its days after an update. 'Repeat\u2026' is withheld. It never marks itself done.", "Both"),
    ("OUR-30", "Ours", "P2", "The quiet wash marks what changed, and never your own edits",
     "Have the other account add and change a few rows. Open the room. Then tick something yourself and reopen.",
     "Their changed rows carry a tint and a firmer border on arrival; a screen reader says 'changed since you last looked'. Your OWN edits are never washed. Leaving and returning clears it.", "Both"),
    ("OUR-31", "Ours", "P2", "Recently removed puts a task back",
     "Remove a shared row. Scroll to the foot of the list.",
     "It sits under 'Recently removed', dimmed, with 'Put it back'. Restoring returns it for both people. Nothing anywhere says which of you removed it.", "Both"),
    ("OUR-32", "Ours", "P2", "Removing a shared REPEAT does not pretend to be 'Skip today'",
     "Hold a repeating shared row and look at the shelf.",
     "It reads 'Remove', not 'Skip today', and the spoken label names removing the repeat. (The shared list has no per-day skip; the pre-fix build promised one and deleted the series for both people.)", "Both"),
    ("OUR-33", "Ours", "P1", "A closed list is still readable, and read-only",
     "Leave a list. Open it from the archive on the pairing screen.",
     "Every row is readable. The capture bar is GONE. Ticking, renaming, removing and re-cadencing all do nothing, and each row says why. Each row still offers taking a copy for yourself.", "Both"),
    ("OUR-34", "Ours", "P2", "The archive lists closed lists, and puts them away",
     "With at least one closed list, look under 'Lists you have kept' on the pairing screen.",
     "Purpose and closed MONTH only, no counts of any kind. 'Put it away' hides it from the default view; 'Show the ones put away' brings it back. Nothing is deleted.", "Both"),
    ("OUR-35", "Ours", "P1", "Reopening together is a handshake, never unilateral",
     "On a closed list, tap 'Reopen together\u2026'. Read the code to the second account and enter it in the ordinary join field.",
     "The offering side gets a code and nothing reopens yet. Entering it in the SAME field used for a new list reopens the list for both, with everything as it was. It is never asked which kind of code it is.", "Both"),
    ("OUR-36", "Ours", "P1", "Delete for good commits when the screen closes, not on a timer",
     "On a closed list, tap 'Delete this list for good'. Read what appears. Tap 'Keep it'. Then do it again and navigate away instead.",
     "The line says nothing has been told to anyone yet and that it deletes when you leave this screen. There is NO countdown anywhere. 'Keep it' cancels it entirely. Leaving the screen with it standing actually deletes it.", "Both"),
    ("OUR-37", "Ours", "P2", "Rename yourself, and only yourself",
     "On a live list, tap the line reading 'They see you as {name}' and type a new name.",
     "Your name changes for both of you. There is no affordance anywhere to edit the OTHER person's name.", "Both"),
    ("OUR-38", "Ours", "P2", "A dropped signal never throws you out of the room",
     "Open the room, then turn off the network and wait for a poll (~15s).",
     "The list stays on screen with a quiet 'no connection' line. You are NOT redirected to the pairing screen, and nothing looks deleted.", "Both"),
    ("OUR-39", "Ours", "P1", "Report this list reaches us and tells nobody else",
     "On a live shared list, scroll to the foot of the pairing screen and tap 'Report this list'. Then check the other account.",
     "The row is quiet, never a red button, and says plainly that only we are told. After tapping it thanks you and says your person has not been told. NOTHING changes on the other account's screen, anywhere. The report email arrives carrying the pair id and no task text.", "Both"),
    ("OUR-40", "Ours", "P1", "A disabled list reads EXACTLY like an ordinary closed one",
     "Set pairs.disabled_at by hand in Supabase for a test pair, then open it on both accounts.",
     "Both see the ordinary 'This list is closed' state, read-only, with the archive row. There is NO mark, word, colour or ordering anywhere that distinguishes it from a list somebody simply left. (Any difference would tell a reported person they were reported.)", "Both"),
    ("OUR-41", "Ours", "P2", "Leaving is the block, and needs no reason",
     "On a live list, tap 'Leave this list'.",
     "It closes immediately for both people, with no confirmation gauntlet, no reason asked, and no way for the other person to reopen it alone. Both can still read everything.", "Both"),
    ("LEG-05", "Legal", "P2", "The shared-list clauses are in both copies of both documents",
     "Read Settings > Privacy and Settings > Terms in-app, then doubledone.app/privacy.html and /terms.html.",
     "Privacy carries the shared-lists sections (what the other person sees, that authorship is not stored, freezing, the 30-day window). Terms carry 'Shared lists, and what is not allowed on them' plus 'Reporting, and what we will do' with the 24-hour aim. In-app and web wording match.", "Both"),
    ("OUR-42", "Ours", "P1", "A change on one phone appears on the other WITHOUT touching it",
     "Both sit on the shared list. One adds a task. Do not touch the other device.",
     "It appears within about 15 seconds. (The poll never fired once before 2026-08-09: setPair rebuilt the pair object every sync, which restarted the interval before it could tick.)", "Both"),
    ("OUR-43", "Ours", "P1", "Opening your own list does not bounce you back",
     "With a live list, tap Ours then 'Open the list'. Also try /ours-list directly in a cold tab.",
     "You land on the list and stay. (useSession returns null while hydrating; the room used to read that as signed-out and redirect, making the button an endless loop.)", "Both"),
    ("OUR-44", "Ours", "P2", "The pairing screen offers a way INTO the list",
     "Finish pairing, then look at the pairing screen.",
     "There is an 'Open the list' button. (It used to dead-end: the only route to the room was Today's door or the Menu, which a new pair has no reason to know about.)", "Both"),
    ("OUR-45", "Ours", "P2", "The quiet wash appears and then takes itself away",
     "Have the other account change a row while you are elsewhere, then open the list.",
     "The changed row carries ONE warmed border, on the card itself, never a second ring around it. It clears itself after about 8 seconds without animating. It does not return on the next visit, even if the other phone's clock runs ahead.", "Both"),
    ("OUR-46", "Ours", "P1", "A repeating task can be ADDED, not only converted",
     "In the shared list, type a title, then tap 'Repeat…' above the box.",
     "The same cadence sheet as everywhere else opens, its commit button names the rhythm, and committing ADDS the task with that cadence. Typing and pressing enter still adds a plain one-off.", "Both"),
    ("OUR-47", "Ours", "P2", "The Ours door on Today is findable",
     "With a live list, look at Today.",
     "A tinted, accented card carrying the list's name, not a hairline footnote. Absent entirely when there is no live list.", "Both"),
    ("OUR-48", "Ours", "P1", "Their tick marks your copy DONE and leaves it on your day",
     "A has an open brought copy on Today. B ticks that row in Ours. A backgrounds the app and returns to Today WITHOUT navigating.",
     "The copy is still there, ticked, NOT gone. It appears in A's Lookback. Nothing anywhere says which of them did it. (Before 2026-08-11 it was tombstoned and replaced by a note, and for three rounds it vanished outright because it was marked done with no completion date.)", "Both"),
    ("OUR-49", "Ours", "P1", "A settled copy carries a completion DATE, not just a done flag",
     "After OUR-48, open Today with ?debug=1 and read the `copy` line for that task.",
     "done=true, doneDay = today's date, showsToday=true. A done task with doneDay=none is invisible on Today AND absent from Lookback: tasksForToday places a finished task by its completion day.", "Both"),
    ("OUR-50", "Ours", "P2", "A finished copy can be brought over again",
     "After OUR-48, hold that row in Ours.",
     "'Bring to my Today' is offered again, not 'Already on your Today'. (A done copy is not on your plate; while only tombstones were skipped you had a task you could neither see nor re-add.)", "Both"),
    ("OUR-51", "Ours", "P1", "Your own edit never highlights; their later edit always does",
     "B ticks a row (A highlights). A un-ticks it. B ticks it again. Watch both screens throughout.",
     "Each side lights up for the OTHER's change every time, including repeatedly on the same row, and never for its own. (`mine` records WHEN you wrote a row, not merely that you did; as a plain id set, one touch made that row permanently dark on that device.)", "Both"),
    ("OUR-52", "Ours", "P2", "The highlight arrives at once and fades away",
     "Have B change a row while A is on the list.",
     "A's row warms immediately, holds about 8 seconds, then fades over roughly 700ms. It does not snap off, does not strobe, and does not return on the next poll for the same change. Reduce-motion gets an instant disappearance instead of a fade.", "Both"),
    ("OUR-53", "Ours", "P2", "Two devices share ONE clock",
     "Open the list on both devices with ?debug=1 and read `skew` on the wash line.",
     "A small value in seconds, not minutes, on both. Every stamp (completion log, last-look, last-write-wins) is compared against server time via server_now(). A device whose clock ran ahead used to stamp its own last-look into the future and could never highlight anything again.", "Both"),
    ("OUR-54", "Ours", "P3", "The debug panel is opt-in and carries no task text",
     "Open /ours-list?debug=1 and /today?debug=1, then the same URLs without the flag.",
     "A small log appears only with the flag, at the top of Today and the foot of the list. Every line is counts, ids and flags. NO task title ever appears in it: these get screenshotted and pasted, and a shared list is two people's words.", "Both"),
    ("OUR-55", "Ours", "P1", "Capture on Ours IS Today's capture",
     "Open a shared list and tap the launcher at the foot of the screen.",
     "The same panel Today opens: the same input, the same Speak button (web), the same door, the same Add button naming its own consequence. Not a lookalike, the same component, so the two can never drift apart.", "Both"),
    ("OUR-56", "Ours", "P1", "Repeating can be set AT capture, without knowing a gesture",
     "In that panel, open the door and choose Daily (or Weekly, or Every N days), then Add.",
     "The button reads 'Add · Daily' before the tap, and the row lands already repeating, visible to both of you. (Before this, the only way to make a repeating shared task was to add it and then somehow know to long-press it. Melroy: 'there is no way to add Repeating Tasks.')", "Both"),
    ("OUR-57", "Ours", "P2", "A dump becomes several shared rows, in the order typed",
     "Type three lines into the Ours capture and Add.",
     "The button reads 'Add 3' and three rows land, in the typed order, on both devices. With a cadence set, all three carry it.", "Both"),
    ("OUR-58", "Ours", "P1", "The shared door offers Repeating ONLY",
     "Open the door on Ours and compare it with the door on Today.",
     "Ours shows one row, Repeating, and its overline says so. NO When row (a shared list is not a day: 'Today' would mean nothing and 'Tomorrow' would be a promise nothing in the room can keep) and NO Steps row (a shared row has no slices field, and breaking a thing down is personal). With nothing set the value line reads 'No repeat', never blank.", "Both"),
    ("OUR-59", "Ours", "P2", "Typed text survives closing the panel",
     "Type half a sentence into the Ours capture, tap Close, tick something, then reopen the panel.",
     "The half sentence is still there. The panel is hidden, never unmounted. The door resets to no-repeat, the words never reset.", "Both"),
    ("OUR-60", "Ours", "P2", "The Ours capture rides above the keyboard",
     "On a real Android device, open the Ours capture and start typing.",
     "The panel and its Add button stay visible above the keyboard. Nothing on this stack does this by itself (edge-to-edge Android ignores softwareKeyboardLayoutMode), so it is done by hand and must be re-checked on a device, never only in a browser.", "Android"),
    ("OUR-61", "Ours", "P3", "No AI on the shared capture",
     "Open the Ours capture with AI enabled in Settings.",
     "No 'Break it down', no 'Sort for me', no 'Scan'. Steps a model proposes would land on a list another person reads, and pointing a model at somebody else's screen is a decision about them. Speak stays: it is on-device dictation, not a server call.", "Both"),
    ("OUR-62", "Ours", "P1", "Scan (premium) reaches the shared capture",
     "Premium, AI enabled in Settings. Open the Ours capture on a phone. Tap Scan and photograph a handwritten shopping list.",
     "The words land IN the capture box, not on the list. You read them, edit them, and Add them yourself. This is the surface Scan most belongs on: the most photographed list anybody owns is the one on the fridge, and that list is shared by definition.", "Android"),
    ("OUR-63", "Ours", "P1", "Scan on Ours meets the paywall, tagged apart from Today's",
     "Sign in as a NON-premium user, open the Ours capture, tap Scan.",
     "The premium screen opens, the same one Today's Scan opens. Nothing is scanned and nothing is charged. The gate logs reason 'ocr_ours', not 'ocr': a paywall met on a surface a second person can see is a different moment from one met alone.", "Both"),
    ("OUR-64", "Ours", "P2", "The photo never reaches the shared list",
     "Scan something into Ours and Add it. Check the other device.",
     "Your person sees the rows you added and nothing else, exactly as if you had typed them. No image is stored anywhere, shared or otherwise: it is sent to be read, then discarded.", "Both"),
    ("OUR-65", "Ours", "P3", "Scan is absent when AI is off",
     "Turn AI off in Settings, then open the Ours capture.",
     "No Scan button, on either screen. Speak stays: it is on-device dictation, not a server call.", "Both"),
    ("OUR-66", "Ours", "P1", "A repeating shared row appears on BOTH Todays on its day",
     "Make a shared task repeat weekly on today's weekday. Open Today on both devices.",
     "It appears on each Today under a 'From Ours' heading, below the day's own tasks. Neither person had to fetch it. This is the answer to 'I do not want to manually decide what gets added to my list for today'.", "Both"),
    ("OUR-67", "Ours", "P1", "An UNDATED shared row never reaches Today",
     "Add an ordinary shared task with the door left on Anytime. Check Today on both devices.",
     "It is in the room and NOWHERE on Today. This is the load-bearing rule: without it, your person adding eight things to the shopping list makes your morning eight heavier without you agreeing to any of it.", "Both"),
    ("OUR-68", "Ours", "P1", "Shared rows never make the day look heavier",
     "With several shared rows showing on Today, check the weight gauge, Plan my day, Lighten, and the close-the-day count.",
     "All of them ignore the shared rows entirely. They are not in `tasks`, so they cannot reach any of it by construction. The day the app promises is finite stays finite.", "Both"),
    ("OUR-69", "Ours", "P1", "Ticking a shared row from Today closes it for both",
     "Tick a row in the From Ours strip on device A. Look at device B's Ours list and Today.",
     "Done on both, within a poll. Nothing anywhere says who did it, and neither room tints it as a stranger's change on the device that did it (noteOursMine).", "Both"),
    ("OUR-70", "Ours", "P2", "A dated shared row shows from its day onward, never before",
     "Add a shared task with the door set to a date two days out. Check Today now, then on the day.",
     "Absent today, present on the day and on any day after (a shared thing nobody did does not stop being one). The room shows its date on the row all along, so it is never a surprise.", "Both"),
    ("OUR-71", "Ours", "P1", "The shared capture rests on Anytime",
     "Open the Ours capture and open the door.",
     "The When row reads Anytime / Today / Tomorrow / Pick a date, with Anytime selected. The Add button says just 'Add'. Choosing Today makes it read 'Add . Today', because that is the notable case: it will appear on both your days.", "Both"),
    ("OUR-72", "Ours", "P2", "A repeat supersedes the day rather than joining it",
     "On the Ours capture, set a Daily repeat with When left on Anytime.",
     "The door summary reads 'Daily', not 'Anytime . Daily'. On Today's own capture it still reads 'Today . Daily', where the when genuinely IS the repeat's start.", "Both"),
    ("OUR-73", "Ours", "P3", "An unreadable cadence is never placed on a day",
     "Have a newer build write a cadence this build cannot parse, then open Today.",
     "It shows in the room (never invisible to the person who set it) and appears on NO Today. This build cannot know which days it means, and a wrong day on a shared surface is a thing the other person has to puzzle over.", "Both"),
    ("OUR-74", "Ours", "P1", "The door says what changed since you looked, and clears when you look",
     "Device B changes something on the shared list. On device A, open Today WITHOUT visiting Ours. Then tap through to Ours and come back.",
     "The Ours door reads '1 since you looked'. After visiting, it is gone. The number is bounded by your own attention, not by the list's length, so it can never sit there climbing at you.", "Both"),
    ("OUR-75", "Ours", "P2", "Your OWN changes never appear in that count",
     "On device A, add or tick several shared rows. Go to Today on the SAME device.",
     "No count on the door. It uses the room's own wash arithmetic, so the door and the room always agree about whose change it was.", "Both"),
    ("OUR-76", "Ours", "P2", "A first-ever visit counts nothing",
     "Pair two accounts, put rows on the list from device B, and open Today on a device A that has never opened the room.",
     "No count. There is no 'last looked' to measure against, and inventing one would open the app with a number already on it.", "Both"),
    ("OUR-77", "Ours", "P1", "The capture's AI note names only what that surface HAS",
     "Type into the Ours capture, then into Today's, and read the line under the buttons.",
     "Ours says Scan sends your photo. Today says Sort for me, Break it down and Scan send what you type or photograph. Ours must NEVER name Sort for me or Break it down: neither button is there, and a false claim about where your data goes teaches people the disclosure is decorative.", "Both"),
    ("OUR-78", "Ours", "P1", "Scan discloses on every path that can reach a camera",
     "Open Scan on web (gallery picker) and on Android with the camera already allowed.",
     "Both show the egress line. It used to appear only on the Android permission screen, so the two paths a person can reach WITHOUT a prompt sent a photograph having said nothing.", "Both"),
    ("OUR-79", "Ours", "P1", "Press and hold a From Ours row on Today",
     "Hold a row in the From Ours strip.",
     "The held card opens with Pin and, on a non-repeating row, Move to. Choosing either TAKES THE ROW ON as your own copy: it leaves the strip and joins your day, where the ordinary reorder and the rest of the held card already work.", "Both"),
    ("OUR-80", "Ours", "P1", "Pin and Move-to never rearrange the other person's day",
     "Move a shared row to tomorrow from your Today. Check the other device.",
     "YOUR copy moves to tomorrow. The shared row is untouched and still on her day. Applied to the shared row itself this would silently take it off her Wednesday too, on a thing she may have planned around.", "Both"),
    ("OUR-81", "Ours", "P2", "A taken-on row is never on screen twice",
     "Take a shared row on (Pin it, or Bring it from the room). Look at Today.",
     "It appears once, as your task, and NOT also in the From Ours strip. Still true once it is ticked: a finished copy and its finished origin must not both render.", "Both"),
    ("OUR-82", "Ours", "P2", "Move-to is not offered on a repeating shared row",
     "Hold a repeating row in the From Ours strip.",
     "Pin is there, Move to is not. A date and a rhythm are alternatives, and rescheduling a series is a question for the room's cadence sheet, not a one-tap action on a day.", "Both"),
    ("OUR-83", "Ours", "P1", "One shared row, one route onto your day",
     "In the room, hold an UNDATED row, then hold a repeating or dated one.",
     "The undated row offers 'Bring to my Today'. The dated or repeating one does NOT: it already arrives by itself. Two routes landing in two different places (a copy in your list, or the strip) was the confusion Melroy hit twice.", "Both"),
    ("OUR-84", "Ours", "P1", "Take a From Ours row on, and it becomes an ordinary task",
     "Hold a row in the From Just Us strip and choose 'Take this on today'.",
     "It leaves the strip and joins your list tagged '. Ours'. Holding it now gives the FULL held card: Break it down, Make tiny, Steps, Nudge, reorder, everything. That visible move IS the message: you have taken it on.", "Both"),
    ("OUR-85", "Ours", "P2", "Break it down works once taken on, and never touches the shared list",
     "Take a shared row on, then Break it down.",
     "The steps land on YOUR copy, on your day. The shared list is unchanged and your person sees nothing new. Breaking down was blocked on the shared row because a model authoring steps onto a list another person reads is a decision about them; on your own copy that objection is gone.", "Both"),
    ("OUR-86", "Ours", "P1", "Select several rows in the room and tick them together",
     "Hold a row, choose Select more, tap a few more, then Done.",
     "All of them tick at once and reach the other device. This is the trip-to-the-shops case: six things bought, six taps was a tax that stops people keeping the list.", "Both"),
    ("OUR-87", "Ours", "P1", "Bulk tick never UN-ticks",
     "Select a mix of done and not-done rows and tap Done.",
     "The not-done ones tick. The already-done ones are untouched. A bulk toggle would silently undo somebody's finished work on a mixed selection.", "Both"),
    ("OUR-88", "Ours", "P1", "Bulk remove warns BEFORE the tap when a repeat is selected",
     "Select a repeating row along with others. Read the shelf.",
     "A line says removing a repeat here ends it for both of you. A shared list has no per-day skip, so this is not a thing to learn from the consequence.", "Both"),
    ("OUR-89", "Ours", "P2", "The select shelf takes the capture bar's seat",
     "Enter select mode in the room.",
     "The capture bar is gone while selecting, and back when you cancel. Two bottom-anchored surfaces competing for the thumb is how a calm screen stops being one, and you are not adding things while you are clearing things.", "Both"),
    ("OUR-90", "Ours", "P3", "No select mode on a closed list",
     "Open an archived list and hold a row.",
     "No Select more. A closed list is readable and nothing else, and the server would refuse the write anyway.", "Both"),
    # --- Check for updates ------------------------------------------------------
    ("UPD-01", "Updates", "P1", "Settings states the version, and up-to-date is a fact",
     "Open Settings and scroll to the foot.",
     "One line: 'v1.2.0 (11)  \u00b7  Up to date'. It is a sage FACT, not a control. Nothing is offered when there is nothing to offer.", "Both"),
    ("UPD-02", "Updates", "P1", "Web offers a reload, and says drafts are kept",
     "With doubledone.app/version.json's \"web\" set ahead of the running build, open Settings on the web.",
     "'A newer version is ready.' + 'Reload to get it' + 'Anything you were typing is kept.' Type something into the capture box first and confirm it survives the reload.", "Web"),
    ("UPD-03", "Updates", "P1", "Native points at the store, and says lists are untouched",
     "With version.json's \"ios\" / \"android\" set ahead of the installed build, open Settings on a device.",
     "'This build is getting old.' + 'Open the store' + 'Your lists stay exactly as they are.' The button opens the correct store listing.", "Android"),
    ("UPD-04", "Updates", "P2", "Unreachable server says NOTHING, never 'up to date'",
     "Put the device in aeroplane mode and open Settings. Then, online, confirm doubledone.app/version.json returns JSON rather than the app's HTML (the SPA fallback would break it silently).",
     "No update line at all, and no 'Up to date' either: it cannot tell, so it says nothing. No spinner and no error.", "Both"),
    ("UPD-05", "Updates", "P2", "The goodnight mention is rare, and framed for the other person",
     "Clear the day with a build two minor versions behind, and reach the rested screen.",
     "At most ONE offer shows. If it is the update one it reads 'The newer one can read everything your person sets, rhythms included.' 'Not now' dismisses it and it does not return for a fortnight. It never displaces the reminder, widget or scrapbook offer.", "Both"),
    ("OUR-12", "Ours", "P3", "Unnamed lists read in each person's language",
     "Create a list WITHOUT naming it. Set the two accounts' devices to different languages (e.g. English and Italian).",
     "Each person sees the app's own word for a shared list in their own language, not the other person's.", "Both"),
]

# A case id is how a tester records a pass or a fail, so two rows sharing one makes the
# result ambiguous and the suite stops being a usable launch gate. This bit twice: a new
# case was appended with the next number in its own block without checking the whole list
# (SEQ-05, SEQ-06, VIS-01 all collided, found 2026-07-25). Fail loudly at generation time
# rather than shipping an .xlsx that reads fine and grades wrong.
_dupes = sorted({c[0] for c in CASES if [x[0] for x in CASES].count(c[0]) > 1})
if _dupes:
    raise SystemExit(
        f"duplicate case ids: {', '.join(_dupes)}\n"
        "Give the NEW case the next free number; leave the existing id alone "
        "(it may already be cited in a past test run)."
    )

HEADERS = ["ID", "Area", "Priority", "Test", "Steps", "Expected result",
           "Platform", "Result", "Findings", "Date"]
WIDTHS = [9, 16, 8, 26, 46, 40, 13, 12, 34, 12]
WRAP_COLS = {4, 5, 6, 9}  # 1-based: Test, Steps, Expected, Findings


def style_header_cell(c):
    c.fill = PatternFill("solid", fgColor=ACCENT)
    c.font = Font(color="FFFFFF", bold=True, size=11)
    c.alignment = Alignment(vertical="center", horizontal="left")


def build_xlsx(path: str) -> None:
    wb = Workbook()

    # --- Intro sheet ---------------------------------------------------------
    intro = wb.active
    intro.title = "Read me"
    intro.sheet_view.showGridLines = False
    intro.column_dimensions["A"].width = 22
    intro.column_dimensions["B"].width = 70

    def line(row, label, value="", bold_label=True, big=False):
        a = intro.cell(row=row, column=1, value=label)
        a.font = Font(bold=bold_label, size=16 if big else 11, color=ACCENT if big else INK)
        a.alignment = Alignment(vertical="top", wrap_text=True)
        b = intro.cell(row=row, column=2, value=value)
        b.font = Font(size=16 if big else 11, color=INK)
        b.alignment = Alignment(vertical="top", wrap_text=True)
        return row + 1

    r = 1
    r = line(r, "DoubleDone", "End-to-end manual test suite", big=True)
    r += 1
    r = line(r, "What this is",
             "The manual QA pass for things only a human on real devices and a real account can verify. "
             "Work the 'Test Suite' tab top to bottom (or filter by Priority). For each row, do the Steps, "
             "compare to the Expected result, set Result from the dropdown, and write what you saw in Findings.")
    r += 1
    r = line(r, "Priorities",
             "P1 = must pass before any public launch.  P2 = important polish.  P3 = edge / nice-to-have.")
    r = line(r, "Result values",
             "Pass  /  Fail  /  Blocked (could not run, e.g. setup missing)  /  Not run.")
    r = line(r, "Tip",
             "Send me the Fails and their Findings and I'll fix them. The Summary tab tallies your results.")
    r += 1
    r = line(r, "— Environment (fill in) —", "")
    r = line(r, "Tester", "")
    r = line(r, "Date started", "")
    r = line(r, "Web URL", "https://doubledone.app")
    r = line(r, "Worker version", "9af9acc1 (or newer)")
    r = line(r, "Commit under test", "c65aa2e (or newer)")
    r = line(r, "Browser + version", "")
    r = line(r, "Android device + OS", "")
    r = line(r, "APK build", "")
    r += 1
    r = line(r, "Note on prerequisites",
             "Some rows need one-time setup first: DEL-00 (create the delete_account function in Supabase), "
             "AUTH-01 (sign in), MCP-00/01 (copy token + connect a client), DEP-02 (sideload the APK). "
             "Do those before the rows that depend on them.")

    # --- Test Suite sheet ----------------------------------------------------
    ws = wb.create_sheet("Test Suite")
    ws.sheet_view.showGridLines = False
    thin = Side(style="thin", color=LINE)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for i, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=i, value=h)
        style_header_cell(c)
        c.border = border
        ws.column_dimensions[get_column_letter(i)].width = WIDTHS[i - 1]

    for ri, case in enumerate(CASES, start=2):
        cid, area, prio, test, steps, expected, platform = case
        values = [cid, area, prio, test, steps, expected, platform, "", "", ""]
        for ci, v in enumerate(values, start=1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.border = border
            c.alignment = Alignment(vertical="top", wrap_text=ci in WRAP_COLS)
            if ci == 1:
                c.font = Font(bold=True, color=INK)
            if ci == 3:  # priority colour
                c.font = Font(bold=True,
                              color={"P1": "9B2D3B", "P2": "8A6D1F", "P3": "6B6B6B"}.get(prio, INK))

    last = len(CASES) + 1

    # Result dropdown + colour rules.
    dv = DataValidation(type="list", formula1='"Pass,Fail,Blocked,Not run"', allow_blank=True)
    dv.error = "Pick Pass, Fail, Blocked or Not run."
    dv.prompt = "Pass / Fail / Blocked / Not run"
    ws.add_data_validation(dv)
    dv.add(f"H2:H{last}")
    for value, fill in (("Pass", PASS_FILL), ("Fail", FAIL_FILL), ("Blocked", BLOCK_FILL)):
        ws.conditional_formatting.add(
            f"H2:H{last}",
            CellIsRule(operator="equal", formula=[f'"{value}"'],
                       fill=PatternFill("solid", fgColor=fill)))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:J{last}"

    # --- Summary sheet -------------------------------------------------------
    s = wb.create_sheet("Summary")
    s.sheet_view.showGridLines = False
    s.column_dimensions["A"].width = 18
    s.column_dimensions["B"].width = 12
    rng = f"'Test Suite'!$H$2:$H${last}"
    s.cell(row=1, column=1, value="Results").font = Font(bold=True, size=14, color=ACCENT)
    rows = [
        ("Total cases", str(len(CASES))),
        ("Pass", f'=COUNTIF({rng},"Pass")'),
        ("Fail", f'=COUNTIF({rng},"Fail")'),
        ("Blocked", f'=COUNTIF({rng},"Blocked")'),
        ("Not run", f'=COUNTIF({rng},"Not run")'),
        ("Unmarked", f'={len(CASES)}-COUNTA({rng})'),
    ]
    for i, (label, val) in enumerate(rows, start=2):
        s.cell(row=i, column=1, value=label).font = Font(bold=label in ("Total cases",), color=INK)
        s.cell(row=i, column=2, value=val)
    s.cell(row=9, column=1, value="P1 cases").font = Font(bold=True, color=INK)
    s.cell(row=9, column=2, value=f'=COUNTIF(\'Test Suite\'!$C$2:$C${last},"P1")')
    s.cell(row=10, column=1, value="P1 passed").font = Font(color=INK)
    s.cell(row=10, column=2,
           value=f'=COUNTIFS(\'Test Suite\'!$C$2:$C${last},"P1",\'Test Suite\'!$H$2:$H${last},"Pass")')

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)


def build_md(path: str) -> None:
    lines = [
        "# DoubleDone, end-to-end test suite",
        "",
        "The readable copy of the manual QA pass. The fillable version with a Result "
        "dropdown is `DoubleDone-E2E-Test-Suite.xlsx` (same content, generated from "
        "`scripts/gen-test-suite.py`).",
        "",
        "**Priorities:** P1 must pass before launch, P2 important polish, P3 edge / nice.",
        "",
    ]
    area = None
    for cid, a, prio, test, steps, expected, platform in CASES:
        if a != area:
            area = a
            lines += ["", f"## {area}", ""]
            lines += ["| ID | Pri | Platform | Test | Steps | Expected |",
                      "|---|---|---|---|---|---|"]
        lines.append(f"| {cid} | {prio} | {platform} | {test} | {steps} | {expected} |")
    lines.append("")
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


if __name__ == "__main__":
    here = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    xlsx = os.path.join(here, "docs", "qa", "DoubleDone-E2E-Test-Suite.xlsx")
    md = os.path.join(here, "docs", "qa", "e2e-test-suite.md")
    build_xlsx(xlsx)
    build_md(md)
    print(f"wrote {xlsx}")
    print(f"wrote {md}")
    print(f"{len(CASES)} cases")
